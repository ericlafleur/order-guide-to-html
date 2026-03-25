"""Main workbook parser orchestrator."""

from collections import OrderedDict
from pathlib import Path
from typing import List

import openpyxl

from ..models.base_models import (
    TrimDef,
    MatrixSheet,
    ColorSheet,
    SpecColumn,
    EngineAxleEntry,
    TraileringRecord,
    GCWRRecord,
    WorkbookData,
)
from .common import parse_filename_metadata
from .matrix_parser import parse_matrix_sheet
from .color_parser import parse_color_sheet
from .spec_parser import parse_spec_sheet
from .engine_axle_parser import parse_engine_axles_sheet
from .trailering_parser import parse_trailering_sheet
from .glossary_parser import parse_glossary_sheet


def parse_workbook(path: Path) -> WorkbookData:
    """Parse complete workbook file.
    
    Args:
        path: Path to Excel workbook
        
    Returns:
        WorkbookData containing all parsed sheets
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    year, make, model, vehicle_name = parse_filename_metadata(path)

    trim_defs: List[TrimDef] = []
    matrix_sheets: List[MatrixSheet] = []
    color_sheets: List[ColorSheet] = []
    spec_columns: List[SpecColumn] = []
    engine_axle_entries: List[EngineAxleEntry] = []
    trailering_records: List[TraileringRecord] = []
    gcwr_records: List[GCWRRecord] = []
    glossary: OrderedDict[str, str] = OrderedDict()

    # Parse each sheet
    for name in wb.sheetnames:
        ws = wb[name]
        
        # Try parsing as matrix sheet
        matrix = parse_matrix_sheet(ws, trim_defs or None)
        if matrix:
            if not trim_defs:
                trim_defs = matrix.trim_defs
            matrix_sheets.append(matrix)
        
        # Parse color sheets
        if name.startswith("Colour and Trim"):
            color_sheets.append(parse_color_sheet(ws))
        
        # Parse spec sheets
        if name.startswith("Dimensions") or name.startswith("Specs") or name in {"Dimensions", "Specs"}:
            spec_columns.extend(parse_spec_sheet(ws))
        
        # Parse engine/axle sheets
        if name.startswith("Engine Axles"):
            engine_axle_entries.extend(parse_engine_axles_sheet(ws))
        
        # Parse trailering sheets
        if name.startswith("Trailering Specs"):
            records, gcwrs = parse_trailering_sheet(ws)
            trailering_records.extend(records)
            gcwr_records.extend(gcwrs)
        
        # Parse glossary
        if name == "All":
            glossary.update(parse_glossary_sheet(ws))

    # Import inference functions here to avoid circular imports
    from ..inference import infer_propulsion, infer_vehicle_type, infer_drive_types
    
    propulsion = infer_propulsion(vehicle_name, engine_axle_entries)
    vehicle_type = infer_vehicle_type(vehicle_name)
    drive_types = infer_drive_types(engine_axle_entries, spec_columns, matrix_sheets, propulsion)

    return WorkbookData(
        path=path,
        year=year,
        make=make,
        model=model,
        vehicle_name=vehicle_name,
        trim_defs=trim_defs,
        matrix_sheets=matrix_sheets,
        color_sheets=color_sheets,
        spec_columns=spec_columns,
        engine_axle_entries=engine_axle_entries,
        trailering_records=trailering_records,
        gcwr_records=gcwr_records,
        glossary=glossary,
        sheet_names=wb.sheetnames,
        propulsion=propulsion,
        vehicle_type=vehicle_type,
        drive_types=drive_types,
    )
