from __future__ import annotations

from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
import re

from .utils import CODE_IN_PARENS_RE, FILLER_TOKENS, FOOTNOTE_LINE_RE, STATUS_LABELS, TRAILING_DIGITS_RE, normalize_text, unique_preserve_order
from .models import ColorExteriorRow, ColorInteriorRow, ColorSheet, EngineAxleEntry, EngineAxleItem, GCWRRecord, MatrixRow, MatrixSheet, SpecCell, SpecColumn, TraileringRecord, TrimDef, WorkbookData


MODEL_CODE_LINE_RE = re.compile(r'^[0-9][A-Z0-9]{4,6}$')
GENERIC_MATRIX_TOP_LABELS = {
    'recommended',
    'custom interior trim and seat combinations',
}


def parse_filename_metadata(path: Path) -> Tuple[str, str, str, str]:
    stem = path.stem
    stem = re.sub(r"\s+Export$", "", stem, flags=re.I)
    stem = re.sub(r"\s+(Retail and Fleet|Retail|Fleet)$", "", stem, flags=re.I)
    parts = stem.split()
    if not parts or not re.fullmatch(r"\d{4}", parts[0]):
        raise ValueError(f"Cannot determine year from filename: {path.name}")
    year = parts[0]
    if len(parts) < 3:
        raise ValueError(f"Cannot determine make/model from filename: {path.name}")
    make = parts[1]
    model_tokens = [p for p in parts[2:] if p not in FILLER_TOKENS]
    if not model_tokens:
        raise ValueError(f"Cannot determine model from filename: {path.name}")
    model = " ".join(model_tokens)
    vehicle_name = f"{year} {make} {model}"
    return year, make, model, vehicle_name

def _sheet_family_from_label(label: object) -> str:
    text = normalize_text(label)
    if not text:
        return ''
    lowered = text.lower()
    if lowered in GENERIC_MATRIX_TOP_LABELS:
        return ''
    return text


def parse_trim_header(value: object, sheet_family: str = '') -> Optional[TrimDef]:
    text = normalize_text(value)
    if not text:
        return None
    lines = [normalize_text(x) for x in text.split("\n") if normalize_text(x)]
    if not lines:
        return None

    family = _sheet_family_from_label(sheet_family)
    if len(lines) == 1:
        return TrimDef(name=lines[0], code=lines[0], raw_header=text, family_label=family)

    code = lines[-1]
    model_code = ''
    name_parts = [lines[0]]

    for mid in lines[1:-1]:
        if MODEL_CODE_LINE_RE.fullmatch(mid):
            if not model_code:
                model_code = mid
        else:
            name_parts.append(mid)

    name = normalize_text(' '.join(name_parts))
    return TrimDef(name=name, code=code, raw_header=text, model_code=model_code, family_label=family)


def find_matrix_header_row(ws) -> Optional[int]:
    for r in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "Description" in values:
            return r
    return None

def parse_footnote_map(text: str) -> Dict[str, str]:
    notes: Dict[str, str] = {}
    for line in normalize_text(text).split("\n"):
        m = FOOTNOTE_LINE_RE.match(line)
        if m:
            notes[m.group(1)] = normalize_text(m.group(2))
    return notes

def split_main_notes_and_bullets(text: str) -> Tuple[str, Dict[str, str], List[str]]:
    lines = [normalize_text(line) for line in normalize_text(text).split("\n") if normalize_text(line)]
    main_lines: List[str] = []
    notes: Dict[str, str] = {}
    bullets: List[str] = []
    in_notes = False
    for line in lines:
        m = FOOTNOTE_LINE_RE.match(line)
        if m:
            notes[m.group(1)] = normalize_text(m.group(2))
            in_notes = True
            continue
        if line.startswith("•"):
            bullets.append(normalize_text(line.lstrip("•").strip()))
            in_notes = True
            continue
        if not in_notes:
            main_lines.append(line)
        else:
            bullets.append(line)
    main = normalize_text(" ".join(main_lines))
    return main, notes, bullets

def parse_status_value(raw: str, row_notes: Dict[str, str], sheet_notes: Dict[str, str]) -> Tuple[str, str, List[str]]:
    raw = normalize_text(raw)
    if not raw:
        return "", "", []
    m = re.match(r"^(--|[A-Z]+|[■□*]+)(.*)$", raw)
    if m:
        code = m.group(1)
        suffix = m.group(2)
    else:
        code = raw
        suffix = ""
    label = STATUS_LABELS.get(code, code)
    note_ids = re.findall(r"\d+", suffix)
    notes: List[str] = []
    for note_id in note_ids:
        if note_id in row_notes:
            notes.append(row_notes[note_id])
        elif note_id in sheet_notes:
            notes.append(sheet_notes[note_id])
    return code, label, unique_preserve_order(notes)

def parse_matrix_sheet(ws, trim_defs: Optional[List[TrimDef]] = None) -> Optional[MatrixSheet]:
    header_row = find_matrix_header_row(ws)
    if header_row is None:
        return None

    headers = [normalize_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    description_col = headers.index("Description") + 1

    legend_text_parts: List[str] = []
    sheet_footnotes: Dict[str, str] = {}
    for r in range(1, header_row):
        row_texts = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        nonempty = [t for t in row_texts if t]
        if nonempty:
            legend_text_parts.extend(nonempty)
        for item in nonempty:
            sheet_footnotes.update(parse_footnote_map(item))

    sheet_family = _sheet_family_from_label(ws.cell(1, 1).value)
    inferred_trim_defs: List[TrimDef] = []
    c = description_col + 1
    while c <= ws.max_column:
        raw = normalize_text(ws.cell(header_row, c).value)
        if not raw:
            break
        parsed = parse_trim_header(raw, sheet_family=sheet_family)
        if parsed:
            inferred_trim_defs.append(parsed)
        c += 1
    active_trim_defs = inferred_trim_defs or trim_defs or []
    trim_cols = list(range(description_col + 1, description_col + 1 + len(active_trim_defs)))

    rows: List[MatrixRow] = []
    current_group: Optional[str] = None
    for r in range(header_row + 1, ws.max_row + 1):
        meta = [normalize_text(ws.cell(r, c).value) for c in range(1, description_col + 1)]
        status_values = [normalize_text(ws.cell(r, c).value) for c in trim_cols]
        if not any(meta) and not any(status_values):
            continue

        if not any(status_values):
            nonempty = [m for m in meta if m]
            for item in nonempty:
                sheet_footnotes.update(parse_footnote_map(item))
            if nonempty and not any(FOOTNOTE_LINE_RE.match(line) for item in nonempty for line in item.split("\n")):
                current_group = nonempty[0]
            continue

        description_raw = meta[-1]
        if not description_raw:
            continue

        option_code = meta[0] or None
        ref_code = meta[1] or None if len(meta) > 1 else None
        aux_meta = [x for x in meta[2:-1] if x]
        main, inline_notes, bullet_notes = split_main_notes_and_bullets(description_raw)

        row = MatrixRow(
            sheet_name=ws.title,
            row_group=current_group,
            option_code=option_code,
            ref_code=ref_code,
            aux_meta=aux_meta,
            description_raw=description_raw,
            description_main=main or description_raw,
            inline_footnotes=inline_notes,
            bullet_notes=bullet_notes,
            status_by_trim={
                active_trim_defs[idx].key: status_values[idx]
                for idx in range(min(len(active_trim_defs), len(status_values)))
                if normalize_text(status_values[idx])
            },
        )
        rows.append(row)

    legend_text = "\n".join(unique_preserve_order(legend_text_parts))
    return MatrixSheet(
        name=ws.title,
        legend_text=legend_text,
        trim_defs=active_trim_defs,
        footnotes=sheet_footnotes,
        rows=rows,
    )

def parse_color_sheet(ws) -> ColorSheet:
    footnotes: Dict[str, str] = {}
    bullet_notes: List[str] = []
    heading_lines: List[str] = []
    interior_rows: List[ColorInteriorRow] = []
    exterior_rows: List[ColorExteriorRow] = []

    for r in range(1, ws.max_row + 1):
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        first = row_values[0] if row_values else ""
        if r <= 3 and any(row_values):
            heading_lines.extend([v for v in row_values if v])
        for value in row_values:
            if value:
                footnotes.update(parse_footnote_map(value))
                for line in normalize_text(value).split("\n"):
                    if line.startswith("•"):
                        bullet_notes.append(normalize_text(line.lstrip("•").strip()))

        if first == "Decor Level":
            color_headers = [normalize_text(ws.cell(r, c).value).split("\n")[0] for c in range(5, ws.max_column + 1)]
            rr = r + 1
            while rr <= ws.max_row:
                values = [normalize_text(ws.cell(rr, c).value) for c in range(1, ws.max_column + 1)]
                if values[0] == "Exterior Solid Paint":
                    break
                if not any(values):
                    rr += 1
                    continue
                if values[0] and values[0] != "Decor Level":
                    colors = {
                        color_headers[i]: values[4 + i]
                        for i in range(len(color_headers))
                        if color_headers[i]
                    }
                    interior_rows.append(
                        ColorInteriorRow(
                            decor_level=values[0],
                            seat_type=values[1],
                            seat_code=values[2],
                            seat_trim=values[3],
                            colors=colors,
                        )
                    )
                rr += 1

        if first == "Exterior Solid Paint":
            color_headers = [normalize_text(ws.cell(r, c).value).split("\n")[0] for c in range(5, ws.max_column + 1)]
            rr = r + 1
            while rr <= ws.max_row:
                values = [normalize_text(ws.cell(rr, c).value) for c in range(1, ws.max_column + 1)]
                if not any(values):
                    rr += 1
                    continue
                if FOOTNOTE_LINE_RE.match(values[0]) or values[0].startswith("•"):
                    break
                if values[0] and values[0] != "Exterior Solid Paint":
                    colors = {
                        color_headers[i]: values[4 + i]
                        for i in range(len(color_headers))
                        if color_headers[i]
                    }
                    exterior_rows.append(
                        ColorExteriorRow(
                            title=values[0],
                            color_code=values[2],
                            touch_up_paint_number=values[3],
                            colors=colors,
                        )
                    )
                rr += 1

    return ColorSheet(
        name=ws.title,
        heading_text=" | ".join(unique_preserve_order(heading_lines)),
        footnotes=footnotes,
        bullet_notes=unique_preserve_order(bullet_notes),
        interior_rows=interior_rows,
        exterior_rows=exterior_rows,
    )

def parse_spec_sheet(ws) -> List[SpecColumn]:
    section_markers = {"Specifications", "Capacities"}
    first_section_row: Optional[int] = None
    for r in range(1, min(ws.max_row, 10) + 1):
        if normalize_text(ws.cell(r, 1).value) in section_markers:
            first_section_row = r
            break
    if first_section_row is None:
        return []

    header_row = first_section_row
    nonempty_on_section_row = sum(1 for c in range(1, ws.max_column + 1) if normalize_text(ws.cell(first_section_row, c).value))
    if nonempty_on_section_row <= 1 and first_section_row > 1:
        header_row = first_section_row - 1

    header_cells = [normalize_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    top_label = ""
    for r in range(1, header_row + 1):
        cell = normalize_text(ws.cell(r, 1).value)
        if cell and cell not in section_markers and "all dimensions" not in cell.lower():
            top_label = cell
            break

    columns: List[SpecColumn] = []
    for c in range(2, ws.max_column + 1):
        header = normalize_text(ws.cell(header_row, c).value)
        if not header:
            continue
        header_lines = [normalize_text(x) for x in header.split("\n") if normalize_text(x)]
        current_section = normalize_text(ws.cell(first_section_row, 1).value) if header_row == first_section_row else ""
        cells: List[SpecCell] = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = normalize_text(ws.cell(r, 1).value)
            row_values = [normalize_text(ws.cell(r, cc).value) for cc in range(1, ws.max_column + 1)]
            if label in section_markers and sum(1 for x in row_values[1:] if x) == 0:
                current_section = label
                continue
            value = normalize_text(ws.cell(r, c).value)
            if label and value and value != "--":
                cells.append(SpecCell(section=current_section or "Data", label=label, value=value))
        if cells:
            columns.append(
                SpecColumn(
                    sheet_name=ws.title,
                    top_label=top_label,
                    header=header,
                    header_lines=header_lines,
                    cells=cells,
                )
            )
    return columns

def parse_engine_axles_sheet(ws) -> List[EngineAxleEntry]:
    if not ws.title.startswith("Engine Axles"):
        return []
    top_label = normalize_text(ws.cell(1, 1).value)
    legend_text = " ".join(
        normalize_text(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 4) + 1)
        for c in range(1, ws.max_column + 1)
        if normalize_text(ws.cell(r, c).value)
    )
    footnotes: Dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        first = normalize_text(ws.cell(r, 1).value)
        if first:
            footnotes.update(parse_footnote_map(first))

    section_headers: Dict[int, str] = {}
    current_section = ""
    for c in range(3, ws.max_column + 1):
        value = normalize_text(ws.cell(3, c).value)
        if value:
            current_section = value
        section_headers[c] = current_section

    entries: List[EngineAxleEntry] = []
    current_model = ""
    for r in range(5, ws.max_row + 1):
        model = normalize_text(ws.cell(r, 1).value)
        engine = normalize_text(ws.cell(r, 2).value)
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not any(row_values):
            continue
        if model and FOOTNOTE_LINE_RE.match(model):
            continue
        if model:
            current_model = model
        if not engine:
            continue
        items: List[EngineAxleItem] = []
        for c in range(3, ws.max_column + 1):
            raw_status = normalize_text(ws.cell(r, c).value)
            if not raw_status or raw_status == "--":
                continue
            code, label, notes = parse_status_value(raw_status, {}, footnotes)
            items.append(
                EngineAxleItem(
                    category=section_headers.get(c, ""),
                    name=normalize_text(ws.cell(4, c).value),
                    raw_status=raw_status,
                    status_code=code,
                    status_label=label,
                    notes=notes,
                )
            )
        if items:
            entries.append(
                EngineAxleEntry(
                    sheet_name=ws.title,
                    top_label=top_label,
                    model_code=current_model,
                    engine=engine,
                    items=items,
                )
            )
    return entries

def parse_value_and_footnote_ids(raw: str) -> Tuple[str, List[str]]:
    raw = normalize_text(raw)
    if not raw or raw == "--":
        return "", []
    m = re.match(r"^(.*\))(\d+)$", raw)
    if m:
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    m = re.match(r"^([0-9]+\.[0-9]{2})(\d+)$", raw)
    if m:
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    m = TRAILING_DIGITS_RE.match(raw)
    if m and re.search(r"[A-Za-z]", m.group(1)):
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    return raw, []

def parse_trailering_sheet(ws) -> Tuple[List[TraileringRecord], List[GCWRRecord]]:
    if not ws.title.startswith("Trailering Specs"):
        return [], []

    sheet_footnotes: Dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            sheet_footnotes.update(parse_footnote_map(normalize_text(ws.cell(r, c).value)))

    rating_note = normalize_text(ws.cell(1, 1).value)
    rating_type = normalize_text(ws.cell(2, 1).value)

    engine_pairs: List[Tuple[str, int, int]] = []
    c = 2
    while c <= ws.max_column:
        engine = normalize_text(ws.cell(3, c).value)
        if engine:
            engine_pairs.append((engine, c, c + 1))
        c += 2

    trailering_records: List[TraileringRecord] = []
    current_model = ""
    gcwr_start_row = None
    for r in range(5, ws.max_row + 1):
        first = normalize_text(ws.cell(r, 1).value)
        if first.startswith("GCWR "):
            gcwr_start_row = r
            break
        if first and FOOTNOTE_LINE_RE.match(first):
            continue
        if first:
            current_model = first
        if not current_model:
            continue
        for engine, axle_col, weight_col in engine_pairs:
            raw_axle = normalize_text(ws.cell(r, axle_col).value)
            raw_weight = normalize_text(ws.cell(r, weight_col).value)
            if (not raw_axle or raw_axle == "--") and (not raw_weight or raw_weight == "--"):
                continue
            axle_ratio, axle_note_ids = parse_value_and_footnote_ids(raw_axle)
            max_weight, weight_note_ids = parse_value_and_footnote_ids(raw_weight)
            note_texts = [sheet_footnotes[nid] for nid in axle_note_ids + weight_note_ids if nid in sheet_footnotes]
            trailering_records.append(
                TraileringRecord(
                    sheet_name=ws.title,
                    rating_type=rating_type,
                    note_text=rating_note,
                    model_code=current_model,
                    engine=engine,
                    axle_ratio=axle_ratio or raw_axle,
                    max_trailer_weight=max_weight or raw_weight,
                    footnotes=unique_preserve_order(note_texts),
                )
            )

    gcwr_records: List[GCWRRecord] = []
    if gcwr_start_row is not None and gcwr_start_row + 3 <= ws.max_row:
        table_title = normalize_text(ws.cell(gcwr_start_row, 1).value)
        header_values = [normalize_text(ws.cell(gcwr_start_row + 2, c).value) for c in range(2, ws.max_column + 1)]
        for r in range(gcwr_start_row + 3, ws.max_row + 1):
            engine = normalize_text(ws.cell(r, 1).value)
            if not engine or FOOTNOTE_LINE_RE.match(engine):
                continue
            for idx, gcwr in enumerate(header_values, start=2):
                if not gcwr:
                    continue
                raw_axle = normalize_text(ws.cell(r, idx).value)
                if not raw_axle or raw_axle == "--":
                    continue
                axle_ratio, note_ids = parse_value_and_footnote_ids(raw_axle)
                gcwr_records.append(
                    GCWRRecord(
                        sheet_name=ws.title,
                        table_title=table_title,
                        engine=engine,
                        gcwr=gcwr,
                        axle_ratio=axle_ratio or raw_axle,
                        footnotes=unique_preserve_order(sheet_footnotes[nid] for nid in note_ids if nid in sheet_footnotes),
                    )
                )

    return trailering_records, gcwr_records

def parse_glossary_sheet(ws) -> OrderedDict[str, str]:
    glossary: OrderedDict[str, str] = OrderedDict()
    first = normalize_text(ws.cell(1, 1).value)
    second = normalize_text(ws.cell(1, 2).value)
    if first != "Option Code" or second != "Description":
        return glossary
    for r in range(2, ws.max_row + 1):
        code = normalize_text(ws.cell(r, 1).value)
        description = normalize_text(ws.cell(r, 2).value)
        if code and description:
            glossary[code] = description
    return glossary

def parse_workbook(path: Path) -> WorkbookData:
    wb = openpyxl.load_workbook(path, data_only=True)
    year, make, model, vehicle_name = parse_filename_metadata(path)

    trim_defs_by_key: OrderedDict[str, TrimDef] = OrderedDict()
    matrix_sheets: List[MatrixSheet] = []
    color_sheets: List[ColorSheet] = []
    spec_columns: List[SpecColumn] = []
    engine_axle_entries: List[EngineAxleEntry] = []
    trailering_records: List[TraileringRecord] = []
    gcwr_records: List[GCWRRecord] = []
    glossary: OrderedDict[str, str] = OrderedDict()

    for name in wb.sheetnames:
        ws = wb[name]
        matrix = parse_matrix_sheet(ws)
        if matrix:
            for trim in matrix.trim_defs:
                trim_defs_by_key.setdefault(trim.key, trim)
            matrix_sheets.append(matrix)
        if name.startswith("Colour and Trim"):
            color_sheets.append(parse_color_sheet(ws))
        if name.startswith("Dimensions") or name.startswith("Specs") or name in {"Dimensions", "Specs"}:
            spec_columns.extend(parse_spec_sheet(ws))
        if name.startswith("Engine Axles"):
            engine_axle_entries.extend(parse_engine_axles_sheet(ws))
        if name.startswith("Trailering Specs"):
            records, gcwrs = parse_trailering_sheet(ws)
            trailering_records.extend(records)
            gcwr_records.extend(gcwrs)
        if name == "All":
            glossary.update(parse_glossary_sheet(ws))

    return WorkbookData(
        path=path,
        year=year,
        make=make,
        model=model,
        vehicle_name=vehicle_name,
        trim_defs=list(trim_defs_by_key.values()),
        matrix_sheets=matrix_sheets,
        color_sheets=color_sheets,
        spec_columns=spec_columns,
        engine_axle_entries=engine_axle_entries,
        trailering_records=trailering_records,
        gcwr_records=gcwr_records,
        glossary=glossary,
        sheet_names=wb.sheetnames,
    )

def _trim_subfamily_name(trim: TrimDef) -> str:
    """Return the sub-model name for a TrimDef.

    For most workbooks, this is just trim.family_label.  For the Corvette
    workbook the 'ZR1 and ZR1X' family_label covers two distinct sub-models;
    this function derives the correct name ('ZR1' or 'ZR1X') from the trim
    body-style name.
    """
    if trim.family_label and trim.family_label != 'ZR1 and ZR1X':
        return trim.family_label
    # Derive from name by stripping the trailing body-style word
    for suffix in (' Coupe', ' Convertible', ' Sedan', ' Hatchback'):
        if trim.name.endswith(suffix):
            return trim.name[:-len(suffix)].strip()
    return trim.name


def _merge_status_values(statuses: List[str]) -> str:
    """Pick the most available status from a list of raw status strings.

    When collapsing body-style variants (e.g. Coupe + Convertible) into a
    single package-level TrimDef, take whichever value indicates the feature
    is most available — 'S' beats 'A' beats '--', etc.  Footnote suffixes
    (e.g. 'S1') are stripped for comparison but preserved in the winner.
    """
    if not statuses:
        return ''
    priority = {'S': 6, '■': 6, '□': 5, 'D': 4, 'A': 3, '--': 1}

    def key(s: str) -> int:
        base = re.sub(r'\d+$', '', s.strip()) if s else ''
        return priority.get(base, priority.get(s, 2))

    return max(statuses, key=key)


def split_workbook_by_subfamily(data: WorkbookData) -> List[WorkbookData]:
    """Partition a multi-family workbook into separate sub-workbooks, one per subfamily.

    For standard single-family workbooks this is a no-op — returns [data].

    For Corvette-style workbooks that bundle multiple distinct sub-models in
    one spreadsheet file:

    - Identifies sub-model families (Stingray, Z06, E-Ray, ZR1, ZR1X) from
      TrimDef.family_label / name prefix.
    - Creates one WorkbookData per sub-model.
    - Collapses body-style variants (Coupe + Convertible) that share a package
      code into a single synthetic TrimDef whose name *is* the package code
      (e.g. '1LT', '2LT', '3LT').  Status values are merged by taking the
      most-available value across body styles.
    - Filters matrix_sheets, spec_columns, engine_axle_entries, and
      trailering_records to those that belong to each sub-model.
    - Colour sheets are shared across sub-workbooks (trim_matches_decor handles
      per-package filtering at render time).
    """
    subfamilies = list(dict.fromkeys(_trim_subfamily_name(t) for t in data.trim_defs))
    # Only split when the workbook genuinely contains multiple distinct sub-models,
    # identified by multiple distinct non-empty family_label values on TrimDefs.
    distinct_family_labels = list(dict.fromkeys(t.family_label for t in data.trim_defs if t.family_label))
    if len(distinct_family_labels) <= 1:
        return [data]

    results: List[WorkbookData] = []

    for sub in subfamilies:
        sub_trims = [t for t in data.trim_defs if _trim_subfamily_name(t) == sub]

        # One collapsed TrimDef per unique package code, in first-seen order
        packages_seen: List[str] = []
        collapsed: List[TrimDef] = []
        for t in sub_trims:
            if t.code not in packages_seen:
                packages_seen.append(t.code)
                collapsed.append(TrimDef(
                    name=t.code,
                    code=t.code,
                    raw_header=f'{sub}\n{t.code}',
                    model_code='',
                    family_label=sub,
                ))

        # package code → original TrimDef keys for status aggregation
        orig_keys_by_code: Dict[str, List[str]] = {}
        for t in sub_trims:
            orig_keys_by_code.setdefault(t.code, []).append(t.key)

        sub_orig_keys = {t.key for t in sub_trims}

        # Rebuild matrix sheets that belong to this sub-model
        new_sheets: List[MatrixSheet] = []
        for sheet in data.matrix_sheets:
            sheet_trim_keys: set = set()
            for row in sheet.rows:
                sheet_trim_keys.update(row.status_by_trim.keys())
            if not (sheet_trim_keys & sub_orig_keys):
                continue

            new_rows: List[MatrixRow] = []
            for row in sheet.rows:
                new_status: Dict[str, str] = {}
                for ct in collapsed:
                    orig_ks = orig_keys_by_code[ct.code]
                    vals = [row.status_by_trim[k] for k in orig_ks if k in row.status_by_trim]
                    if vals:
                        new_status[ct.key] = _merge_status_values(vals)
                if new_status:
                    new_rows.append(MatrixRow(
                        sheet_name=row.sheet_name,
                        row_group=row.row_group,
                        option_code=row.option_code,
                        ref_code=row.ref_code,
                        aux_meta=list(row.aux_meta),
                        description_raw=row.description_raw,
                        description_main=row.description_main,
                        inline_footnotes=dict(row.inline_footnotes),
                        bullet_notes=list(row.bullet_notes),
                        status_by_trim=new_status,
                    ))
            if new_rows:
                new_sheets.append(MatrixSheet(
                    name=sheet.name,
                    legend_text=sheet.legend_text,
                    trim_defs=collapsed,
                    footnotes=dict(sheet.footnotes),
                    rows=new_rows,
                ))

        # Filter spec_columns to this sub-model by model code or body-style name
        sub_model_codes = {t.model_code for t in sub_trims if t.model_code}
        sub_body_names = {t.name.lower() for t in sub_trims}

        def col_belongs(col: SpecColumn) -> bool:
            text = ' '.join([col.header, col.top_label] + col.header_lines).lower()
            if sub_model_codes:
                return any(mc.lower() in text for mc in sub_model_codes)
            # Fallback when no model codes available: match by body-style name
            return any(bn in text for bn in sub_body_names)

        sub_spec_cols = [c for c in data.spec_columns if col_belongs(c)]

        # Filter engine_axle_entries and trailering_records by model code
        sub_engine_axles = [
            e for e in data.engine_axle_entries
            if not sub_model_codes or not e.model_code or e.model_code in sub_model_codes
        ]
        sub_trailering = [
            r for r in data.trailering_records
            if not sub_model_codes or not r.model_code or r.model_code in sub_model_codes
        ]

        sub_model = f'{data.model} {sub}'
        sub_vehicle_name = f'{data.year} {data.make} {sub_model}'
        sub_path = data.path.parent / (
            data.path.stem + '_' + sub.replace('-', '').replace(' ', '_') + '.xlsx'
        )
        sub_sheet_names = unique_preserve_order(
            [s.name for s in new_sheets]
            + [c.sheet_name for c in sub_spec_cols]
            + [s.name for s in data.color_sheets]
        )

        results.append(WorkbookData(
            path=sub_path,
            year=data.year,
            make=data.make,
            model=sub_model,
            vehicle_name=sub_vehicle_name,
            trim_defs=collapsed,
            matrix_sheets=new_sheets,
            color_sheets=list(data.color_sheets),
            spec_columns=sub_spec_cols,
            engine_axle_entries=sub_engine_axles,
            trailering_records=sub_trailering,
            gcwr_records=list(data.gcwr_records),
            glossary=data.glossary,
            sheet_names=sub_sheet_names,
        ))

    return results

def referenced_codes_for_text(text: str, glossary: Dict[str, str]) -> List[Tuple[str, str]]:
    codes = []
    for code in CODE_IN_PARENS_RE.findall(text):
        if code in glossary:
            codes.append((code, glossary[code]))
    return list(OrderedDict(((code, desc), None) for code, desc in codes).keys())
