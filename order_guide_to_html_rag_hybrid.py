
#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import re
import sys
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


STATUS_LABELS = OrderedDict(
    [
        ("S", "Standard Equipment"),
        ("A", "Available"),
        ("D", "ADI Available"),
        ("■", "Included in Equipment Group"),
        ("□", "Included in Equipment Group but upgradeable"),
        ("*", "Indicates availability of feature on multiple models"),
        ("--", "Not Available"),
    ]
)

FOOTNOTE_LINE_RE = re.compile(r"^\s*(\d+)\.\s*(.+?)\s*$")
URL_RE = re.compile(r"(https?://[^\s<]+)")
TRAILING_DIGITS_RE = re.compile(r"^(.*?)(\d+)\s*$")
CODE_IN_PARENS_RE = re.compile(r"\(([A-Z0-9]{2,6})\)")
NON_ALNUM_RE = re.compile(r"[^a-z0-9]+", re.I)
FILLER_TOKENS = {"Truck", "Trucks", "Cars", "Car", "SUV", "SUVs"}


@dataclass
class TrimDef:
    name: str
    code: str
    raw_header: str

    @property
    def key(self) -> str:
        return self.code or self.name

    @property
    def label(self) -> str:
        if self.code and self.name:
            return f"{self.name} ({self.code})"
        return self.name or self.code


@dataclass
class MatrixRow:
    sheet_name: str
    row_group: Optional[str]
    option_code: Optional[str]
    ref_code: Optional[str]
    aux_meta: List[str]
    description_raw: str
    description_main: str
    inline_footnotes: Dict[str, str] = field(default_factory=dict)
    bullet_notes: List[str] = field(default_factory=list)
    status_by_trim: Dict[str, str] = field(default_factory=dict)

    @property
    def label(self) -> str:
        text = self.description_main or self.description_raw
        text = normalize_text(text).split("\n")[0]
        text = re.sub(r"^NEW!\s+", "", text, flags=re.I)
        short = text
        if ", includes " in text.lower():
            short = text.split(",", 1)[0]
        elif ". " in text and len(text.split(". ", 1)[0]) >= 12:
            short = text.split(". ", 1)[0]
        if len(short) > 140:
            short = short[:137].rstrip() + "..."
        return normalize_text(short)

    @property
    def identity_key(self) -> str:
        basis = " | ".join(
            normalize_text(x)
            for x in [self.option_code or "", self.ref_code or "", self.description_main or self.description_raw]
            if normalize_text(x)
        )
        return NON_ALNUM_RE.sub(" ", basis).strip().lower()


@dataclass
class MatrixSheet:
    name: str
    legend_text: str
    trim_defs: List[TrimDef]
    footnotes: Dict[str, str]
    rows: List[MatrixRow]


@dataclass
class ColorInteriorRow:
    decor_level: str
    seat_type: str
    seat_code: str
    seat_trim: str
    colors: Dict[str, str]


@dataclass
class ColorExteriorRow:
    title: str
    color_code: str
    touch_up_paint_number: str
    colors: Dict[str, str]


@dataclass
class ColorSheet:
    name: str
    heading_text: str
    footnotes: Dict[str, str]
    bullet_notes: List[str]
    interior_rows: List[ColorInteriorRow]
    exterior_rows: List[ColorExteriorRow]


@dataclass
class SpecCell:
    section: str
    label: str
    value: str


@dataclass
class SpecColumn:
    sheet_name: str
    top_label: str
    header: str
    header_lines: List[str]
    cells: List[SpecCell]


@dataclass
class EngineAxleItem:
    category: str
    name: str
    raw_status: str
    status_code: str
    status_label: str
    notes: List[str]


@dataclass
class EngineAxleEntry:
    sheet_name: str
    top_label: str
    model_code: str
    engine: str
    items: List[EngineAxleItem]


@dataclass
class TraileringRecord:
    sheet_name: str
    rating_type: str
    note_text: str
    model_code: str
    engine: str
    axle_ratio: str
    max_trailer_weight: str
    footnotes: List[str]


@dataclass
class GCWRRecord:
    sheet_name: str
    table_title: str
    engine: str
    gcwr: str
    axle_ratio: str
    footnotes: List[str]


@dataclass
class WorkbookData:
    path: Path
    year: str
    make: str
    model: str
    vehicle_name: str
    trim_defs: List[TrimDef]
    matrix_sheets: List[MatrixSheet]
    color_sheets: List[ColorSheet]
    spec_columns: List[SpecColumn]
    engine_axle_entries: List[EngineAxleEntry]
    trailering_records: List[TraileringRecord]
    gcwr_records: List[GCWRRecord]
    glossary: OrderedDict[str, str]
    sheet_names: List[str]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slugify(text: str) -> str:
    text = normalize_text(text).replace("/", " ")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def htmlize_text(text: str) -> str:
    escaped = html.escape(normalize_text(text))
    return URL_RE.sub(r'<a href="\1" target="_blank">\1</a>', escaped)


def parse_filename_metadata(path: Path) -> Tuple[str, str, str, str]:
    stem = path.stem
    stem = re.sub(r"\s+Export$", "", stem, flags=re.I)
    stem = re.sub(r"\s+(Retail and Fleet|Retail|Fleet)$", "", stem, flags=re.I)
    parts = stem.split()
    if not parts or not re.fullmatch(r"\d{4}", parts[0]):
        raise ValueError(f"Cannot determine year from filename: {path.name}")
    year = parts[0]
    if len(parts) < 3:
        raise ValueError(f"Cannot determine make/model from filename: {path.name}")
    make = parts[1]
    model_tokens = [p for p in parts[2:] if p not in FILLER_TOKENS]
    if not model_tokens:
        raise ValueError(f"Cannot determine model from filename: {path.name}")
    model = " ".join(model_tokens)
    vehicle_name = f"{year} {make} {model}"
    return year, make, model, vehicle_name


def parse_trim_header(value: object) -> Optional[TrimDef]:
    text = normalize_text(value)
    if not text:
        return None
    lines = [normalize_text(x) for x in text.split("\n") if normalize_text(x)]
    if not lines:
        return None
    if len(lines) == 1:
        return TrimDef(name=lines[0], code=lines[0], raw_header=text)
    return TrimDef(name=" ".join(lines[:-1]), code=lines[-1], raw_header=text)


def find_matrix_header_row(ws) -> Optional[int]:
    for r in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "Description" in values:
            return r
    return None


def parse_footnote_map(text: str) -> Dict[str, str]:
    notes: Dict[str, str] = {}
    for line in normalize_text(text).split("\n"):
        m = FOOTNOTE_LINE_RE.match(line)
        if m:
            notes[m.group(1)] = normalize_text(m.group(2))
    return notes


def split_main_notes_and_bullets(text: str) -> Tuple[str, Dict[str, str], List[str]]:
    lines = [normalize_text(line) for line in normalize_text(text).split("\n") if normalize_text(line)]
    main_lines: List[str] = []
    notes: Dict[str, str] = {}
    bullets: List[str] = []
    in_notes = False
    for line in lines:
        m = FOOTNOTE_LINE_RE.match(line)
        if m:
            notes[m.group(1)] = normalize_text(m.group(2))
            in_notes = True
            continue
        if line.startswith("•"):
            bullets.append(normalize_text(line.lstrip("•").strip()))
            in_notes = True
            continue
        if not in_notes:
            main_lines.append(line)
        else:
            bullets.append(line)
    main = normalize_text(" ".join(main_lines))
    return main, notes, bullets


def parse_status_value(raw: str, row_notes: Dict[str, str], sheet_notes: Dict[str, str]) -> Tuple[str, str, List[str]]:
    raw = normalize_text(raw)
    if not raw:
        return "", "", []
    m = re.match(r"^(--|[A-Z]+|[■□*]+)(.*)$", raw)
    if m:
        code = m.group(1)
        suffix = m.group(2)
    else:
        code = raw
        suffix = ""
    label = STATUS_LABELS.get(code, code)
    note_ids = re.findall(r"\d+", suffix)
    notes: List[str] = []
    for note_id in note_ids:
        if note_id in row_notes:
            notes.append(row_notes[note_id])
        elif note_id in sheet_notes:
            notes.append(sheet_notes[note_id])
    return code, label, unique_preserve_order(notes)


def unique_preserve_order(items: Iterable[str]) -> List[str]:
    seen = OrderedDict()
    for item in items:
        item = normalize_text(item)
        if item:
            seen[item] = None
    return list(seen.keys())


def parse_matrix_sheet(ws, trim_defs: Optional[List[TrimDef]] = None) -> Optional[MatrixSheet]:
    header_row = find_matrix_header_row(ws)
    if header_row is None:
        return None

    headers = [normalize_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    description_col = headers.index("Description") + 1

    legend_text_parts: List[str] = []
    sheet_footnotes: Dict[str, str] = {}
    for r in range(1, header_row):
        row_texts = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        nonempty = [t for t in row_texts if t]
        if nonempty:
            legend_text_parts.extend(nonempty)
        for item in nonempty:
            sheet_footnotes.update(parse_footnote_map(item))

    inferred_trim_defs: List[TrimDef] = []
    c = description_col + 1
    while c <= ws.max_column:
        raw = normalize_text(ws.cell(header_row, c).value)
        if not raw:
            break
        parsed = parse_trim_header(raw)
        if parsed:
            inferred_trim_defs.append(parsed)
        c += 1
    active_trim_defs = trim_defs or inferred_trim_defs
    trim_cols = list(range(description_col + 1, description_col + 1 + len(active_trim_defs)))

    rows: List[MatrixRow] = []
    current_group: Optional[str] = None
    for r in range(header_row + 1, ws.max_row + 1):
        meta = [normalize_text(ws.cell(r, c).value) for c in range(1, description_col + 1)]
        status_values = [normalize_text(ws.cell(r, c).value) for c in trim_cols]
        if not any(meta) and not any(status_values):
            continue

        if not any(status_values):
            nonempty = [m for m in meta if m]
            for item in nonempty:
                sheet_footnotes.update(parse_footnote_map(item))
            if nonempty and not any(FOOTNOTE_LINE_RE.match(line) for item in nonempty for line in item.split("\n")):
                current_group = nonempty[0]
            continue

        description_raw = meta[-1]
        if not description_raw:
            continue

        option_code = meta[0] or None
        ref_code = meta[1] or None if len(meta) > 1 else None
        aux_meta = [x for x in meta[2:-1] if x]
        main, inline_notes, bullet_notes = split_main_notes_and_bullets(description_raw)

        row = MatrixRow(
            sheet_name=ws.title,
            row_group=current_group,
            option_code=option_code,
            ref_code=ref_code,
            aux_meta=aux_meta,
            description_raw=description_raw,
            description_main=main or description_raw,
            inline_footnotes=inline_notes,
            bullet_notes=bullet_notes,
            status_by_trim={
                active_trim_defs[idx].key: status_values[idx]
                for idx in range(min(len(active_trim_defs), len(status_values)))
                if normalize_text(status_values[idx])
            },
        )
        rows.append(row)

    legend_text = "\n".join(unique_preserve_order(legend_text_parts))
    return MatrixSheet(
        name=ws.title,
        legend_text=legend_text,
        trim_defs=active_trim_defs,
        footnotes=sheet_footnotes,
        rows=rows,
    )


def parse_color_sheet(ws) -> ColorSheet:
    footnotes: Dict[str, str] = {}
    bullet_notes: List[str] = []
    heading_lines: List[str] = []
    interior_rows: List[ColorInteriorRow] = []
    exterior_rows: List[ColorExteriorRow] = []

    for r in range(1, ws.max_row + 1):
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        first = row_values[0] if row_values else ""
        if r <= 3 and any(row_values):
            heading_lines.extend([v for v in row_values if v])
        for value in row_values:
            if value:
                footnotes.update(parse_footnote_map(value))
                for line in normalize_text(value).split("\n"):
                    if line.startswith("•"):
                        bullet_notes.append(normalize_text(line.lstrip("•").strip()))

        if first == "Decor Level":
            color_headers = [normalize_text(ws.cell(r, c).value).split("\n")[0] for c in range(5, ws.max_column + 1)]
            rr = r + 1
            while rr <= ws.max_row:
                values = [normalize_text(ws.cell(rr, c).value) for c in range(1, ws.max_column + 1)]
                if values[0] == "Exterior Solid Paint":
                    break
                if not any(values):
                    rr += 1
                    continue
                if values[0] and values[0] != "Decor Level":
                    colors = {
                        color_headers[i]: values[4 + i]
                        for i in range(len(color_headers))
                        if color_headers[i]
                    }
                    interior_rows.append(
                        ColorInteriorRow(
                            decor_level=values[0],
                            seat_type=values[1],
                            seat_code=values[2],
                            seat_trim=values[3],
                            colors=colors,
                        )
                    )
                rr += 1

        if first == "Exterior Solid Paint":
            color_headers = [normalize_text(ws.cell(r, c).value).split("\n")[0] for c in range(5, ws.max_column + 1)]
            rr = r + 1
            while rr <= ws.max_row:
                values = [normalize_text(ws.cell(rr, c).value) for c in range(1, ws.max_column + 1)]
                if not any(values):
                    rr += 1
                    continue
                if FOOTNOTE_LINE_RE.match(values[0]) or values[0].startswith("•"):
                    break
                if values[0] and values[0] != "Exterior Solid Paint":
                    colors = {
                        color_headers[i]: values[4 + i]
                        for i in range(len(color_headers))
                        if color_headers[i]
                    }
                    exterior_rows.append(
                        ColorExteriorRow(
                            title=values[0],
                            color_code=values[2],
                            touch_up_paint_number=values[3],
                            colors=colors,
                        )
                    )
                rr += 1

    return ColorSheet(
        name=ws.title,
        heading_text=" | ".join(unique_preserve_order(heading_lines)),
        footnotes=footnotes,
        bullet_notes=unique_preserve_order(bullet_notes),
        interior_rows=interior_rows,
        exterior_rows=exterior_rows,
    )


def parse_spec_sheet(ws) -> List[SpecColumn]:
    section_markers = {"Specifications", "Capacities"}
    first_section_row: Optional[int] = None
    for r in range(1, min(ws.max_row, 10) + 1):
        if normalize_text(ws.cell(r, 1).value) in section_markers:
            first_section_row = r
            break
    if first_section_row is None:
        return []

    header_row = first_section_row
    nonempty_on_section_row = sum(1 for c in range(1, ws.max_column + 1) if normalize_text(ws.cell(first_section_row, c).value))
    if nonempty_on_section_row <= 1 and first_section_row > 1:
        header_row = first_section_row - 1

    header_cells = [normalize_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    top_label = ""
    for r in range(1, header_row + 1):
        cell = normalize_text(ws.cell(r, 1).value)
        if cell and cell not in section_markers and "all dimensions" not in cell.lower():
            top_label = cell
            break

    columns: List[SpecColumn] = []
    for c in range(2, ws.max_column + 1):
        header = normalize_text(ws.cell(header_row, c).value)
        if not header:
            continue
        header_lines = [normalize_text(x) for x in header.split("\n") if normalize_text(x)]
        current_section = normalize_text(ws.cell(first_section_row, 1).value) if header_row == first_section_row else ""
        cells: List[SpecCell] = []
        for r in range(header_row + 1, ws.max_row + 1):
            label = normalize_text(ws.cell(r, 1).value)
            row_values = [normalize_text(ws.cell(r, cc).value) for cc in range(1, ws.max_column + 1)]
            if label in section_markers and sum(1 for x in row_values[1:] if x) == 0:
                current_section = label
                continue
            value = normalize_text(ws.cell(r, c).value)
            if label and value and value != "--":
                cells.append(SpecCell(section=current_section or "Data", label=label, value=value))
        if cells:
            columns.append(
                SpecColumn(
                    sheet_name=ws.title,
                    top_label=top_label,
                    header=header,
                    header_lines=header_lines,
                    cells=cells,
                )
            )
    return columns


def parse_engine_axles_sheet(ws) -> List[EngineAxleEntry]:
    if not ws.title.startswith("Engine Axles"):
        return []
    top_label = normalize_text(ws.cell(1, 1).value)
    legend_text = " ".join(
        normalize_text(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 4) + 1)
        for c in range(1, ws.max_column + 1)
        if normalize_text(ws.cell(r, c).value)
    )
    footnotes: Dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        first = normalize_text(ws.cell(r, 1).value)
        if first:
            footnotes.update(parse_footnote_map(first))

    section_headers: Dict[int, str] = {}
    current_section = ""
    for c in range(3, ws.max_column + 1):
        value = normalize_text(ws.cell(3, c).value)
        if value:
            current_section = value
        section_headers[c] = current_section

    entries: List[EngineAxleEntry] = []
    current_model = ""
    for r in range(5, ws.max_row + 1):
        model = normalize_text(ws.cell(r, 1).value)
        engine = normalize_text(ws.cell(r, 2).value)
        row_values = [normalize_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if not any(row_values):
            continue
        if model and FOOTNOTE_LINE_RE.match(model):
            continue
        if model:
            current_model = model
        if not engine:
            continue
        items: List[EngineAxleItem] = []
        for c in range(3, ws.max_column + 1):
            raw_status = normalize_text(ws.cell(r, c).value)
            if not raw_status or raw_status == "--":
                continue
            code, label, notes = parse_status_value(raw_status, {}, footnotes)
            items.append(
                EngineAxleItem(
                    category=section_headers.get(c, ""),
                    name=normalize_text(ws.cell(4, c).value),
                    raw_status=raw_status,
                    status_code=code,
                    status_label=label,
                    notes=notes,
                )
            )
        if items:
            entries.append(
                EngineAxleEntry(
                    sheet_name=ws.title,
                    top_label=top_label,
                    model_code=current_model,
                    engine=engine,
                    items=items,
                )
            )
    return entries


def parse_value_and_footnote_ids(raw: str) -> Tuple[str, List[str]]:
    raw = normalize_text(raw)
    if not raw or raw == "--":
        return "", []
    m = re.match(r"^(.*\))(\d+)$", raw)
    if m:
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    m = re.match(r"^([0-9]+\.[0-9]{2})(\d+)$", raw)
    if m:
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    m = TRAILING_DIGITS_RE.match(raw)
    if m and re.search(r"[A-Za-z]", m.group(1)):
        return normalize_text(m.group(1)), re.findall(r"\d+", m.group(2))
    return raw, []


def parse_trailering_sheet(ws) -> Tuple[List[TraileringRecord], List[GCWRRecord]]:
    if not ws.title.startswith("Trailering Specs"):
        return [], []

    sheet_footnotes: Dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            sheet_footnotes.update(parse_footnote_map(normalize_text(ws.cell(r, c).value)))

    rating_note = normalize_text(ws.cell(1, 1).value)
    rating_type = normalize_text(ws.cell(2, 1).value)

    engine_pairs: List[Tuple[str, int, int]] = []
    c = 2
    while c <= ws.max_column:
        engine = normalize_text(ws.cell(3, c).value)
        if engine:
            engine_pairs.append((engine, c, c + 1))
        c += 2

    trailering_records: List[TraileringRecord] = []
    current_model = ""
    gcwr_start_row = None
    for r in range(5, ws.max_row + 1):
        first = normalize_text(ws.cell(r, 1).value)
        if first.startswith("GCWR "):
            gcwr_start_row = r
            break
        if first and FOOTNOTE_LINE_RE.match(first):
            continue
        if first:
            current_model = first
        if not current_model:
            continue
        for engine, axle_col, weight_col in engine_pairs:
            raw_axle = normalize_text(ws.cell(r, axle_col).value)
            raw_weight = normalize_text(ws.cell(r, weight_col).value)
            if (not raw_axle or raw_axle == "--") and (not raw_weight or raw_weight == "--"):
                continue
            axle_ratio, axle_note_ids = parse_value_and_footnote_ids(raw_axle)
            max_weight, weight_note_ids = parse_value_and_footnote_ids(raw_weight)
            note_texts = [sheet_footnotes[nid] for nid in axle_note_ids + weight_note_ids if nid in sheet_footnotes]
            trailering_records.append(
                TraileringRecord(
                    sheet_name=ws.title,
                    rating_type=rating_type,
                    note_text=rating_note,
                    model_code=current_model,
                    engine=engine,
                    axle_ratio=axle_ratio or raw_axle,
                    max_trailer_weight=max_weight or raw_weight,
                    footnotes=unique_preserve_order(note_texts),
                )
            )

    gcwr_records: List[GCWRRecord] = []
    if gcwr_start_row is not None and gcwr_start_row + 3 <= ws.max_row:
        table_title = normalize_text(ws.cell(gcwr_start_row, 1).value)
        header_values = [normalize_text(ws.cell(gcwr_start_row + 2, c).value) for c in range(2, ws.max_column + 1)]
        for r in range(gcwr_start_row + 3, ws.max_row + 1):
            engine = normalize_text(ws.cell(r, 1).value)
            if not engine or FOOTNOTE_LINE_RE.match(engine):
                continue
            for idx, gcwr in enumerate(header_values, start=2):
                if not gcwr:
                    continue
                raw_axle = normalize_text(ws.cell(r, idx).value)
                if not raw_axle or raw_axle == "--":
                    continue
                axle_ratio, note_ids = parse_value_and_footnote_ids(raw_axle)
                gcwr_records.append(
                    GCWRRecord(
                        sheet_name=ws.title,
                        table_title=table_title,
                        engine=engine,
                        gcwr=gcwr,
                        axle_ratio=axle_ratio or raw_axle,
                        footnotes=unique_preserve_order(sheet_footnotes[nid] for nid in note_ids if nid in sheet_footnotes),
                    )
                )

    return trailering_records, gcwr_records


def parse_glossary_sheet(ws) -> OrderedDict[str, str]:
    glossary: OrderedDict[str, str] = OrderedDict()
    first = normalize_text(ws.cell(1, 1).value)
    second = normalize_text(ws.cell(1, 2).value)
    if first != "Option Code" or second != "Description":
        return glossary
    for r in range(2, ws.max_row + 1):
        code = normalize_text(ws.cell(r, 1).value)
        description = normalize_text(ws.cell(r, 2).value)
        if code and description:
            glossary[code] = description
    return glossary


def parse_workbook(path: Path) -> WorkbookData:
    wb = openpyxl.load_workbook(path, data_only=True)
    year, make, model, vehicle_name = parse_filename_metadata(path)

    trim_defs: List[TrimDef] = []
    matrix_sheets: List[MatrixSheet] = []
    color_sheets: List[ColorSheet] = []
    spec_columns: List[SpecColumn] = []
    engine_axle_entries: List[EngineAxleEntry] = []
    trailering_records: List[TraileringRecord] = []
    gcwr_records: List[GCWRRecord] = []
    glossary: OrderedDict[str, str] = OrderedDict()

    for name in wb.sheetnames:
        ws = wb[name]
        matrix = parse_matrix_sheet(ws, trim_defs or None)
        if matrix:
            if not trim_defs:
                trim_defs = matrix.trim_defs
            matrix_sheets.append(matrix)
        if name.startswith("Colour and Trim"):
            color_sheets.append(parse_color_sheet(ws))
        if name.startswith("Dimensions") or name.startswith("Specs") or name in {"Dimensions", "Specs"}:
            spec_columns.extend(parse_spec_sheet(ws))
        if name.startswith("Engine Axles"):
            engine_axle_entries.extend(parse_engine_axles_sheet(ws))
        if name.startswith("Trailering Specs"):
            records, gcwrs = parse_trailering_sheet(ws)
            trailering_records.extend(records)
            gcwr_records.extend(gcwrs)
        if name == "All":
            glossary.update(parse_glossary_sheet(ws))

    return WorkbookData(
        path=path,
        year=year,
        make=make,
        model=model,
        vehicle_name=vehicle_name,
        trim_defs=trim_defs,
        matrix_sheets=matrix_sheets,
        color_sheets=color_sheets,
        spec_columns=spec_columns,
        engine_axle_entries=engine_axle_entries,
        trailering_records=trailering_records,
        gcwr_records=gcwr_records,
        glossary=glossary,
        sheet_names=wb.sheetnames,
    )



def referenced_codes_for_text(text: str, glossary: Dict[str, str]) -> List[Tuple[str, str]]:
    codes = []
    for code in CODE_IN_PARENS_RE.findall(text):
        if code in glossary:
            codes.append((code, glossary[code]))
    return list(OrderedDict(((code, desc), None) for code, desc in codes).keys())


@dataclass
class ModelFeatureAggregate:
    title: str
    description: str
    orderable_code: str
    reference_code: str
    source_contexts: List[str] = field(default_factory=list)
    availability_contexts: OrderedDict[Tuple[Tuple[Tuple[str, str, Tuple[str, ...]], ...]], List[str]] = field(default_factory=OrderedDict)
    notes: List[str] = field(default_factory=list)
    referenced_codes: List[Tuple[str, str]] = field(default_factory=list)


@dataclass
class TrimFeatureAggregate:
    title: str
    description: str
    orderable_code: str
    reference_code: str
    availability_contexts: OrderedDict[Tuple[str, str], List[str]] = field(default_factory=OrderedDict)
    source_contexts: List[str] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)
    referenced_codes: List[Tuple[str, str]] = field(default_factory=list)


STATUS_PRIORITY = {
    'Standard Equipment': 0,
    'Included in Equipment Group': 1,
    'Included in Equipment Group but upgradeable': 2,
    'ADI Available': 3,
    'Available': 4,
    'Indicates availability of feature on multiple models': 5,
    'Not Available': 6,
}


def feature_title(label: str, orderable_code: str = '', reference_code: str = '') -> str:
    prefix = orderable_code or reference_code
    label = normalize_text(label)
    if prefix:
        return f'{prefix} | {label}'
    return label


def source_context(sheet_name: str, row_group: Optional[str] = None) -> str:
    sheet_name = normalize_text(sheet_name)
    row_group = normalize_text(row_group)
    if row_group and row_group.lower() != sheet_name.lower():
        return f'{sheet_name} | {row_group}'
    return sheet_name


def sentence_chunks(text: str, max_words: int = 110) -> List[str]:
    text = normalize_text(text)
    if not text:
        return []
    parts = [normalize_text(p) for p in re.split(r'(?<=[.!?])\s+|\n+', text) if normalize_text(p)]
    if not parts:
        return [text]
    chunks: List[str] = []
    current: List[str] = []
    current_words = 0
    for part in parts:
        words = len(part.split())
        if words > max_words and not current:
            raw_words = part.split()
            for i in range(0, len(raw_words), max_words):
                chunks.append(' '.join(raw_words[i:i + max_words]))
            continue
        if current and current_words + words > max_words:
            chunks.append(normalize_text(' '.join(current)))
            current = [part]
            current_words = words
        else:
            current.append(part)
            current_words += words
    if current:
        chunks.append(normalize_text(' '.join(current)))
    return [c for c in chunks if c]


def chunk_list(items: Sequence[str], max_words: int = 110, max_items: int = 8) -> List[List[str]]:
    chunks: List[List[str]] = []
    current: List[str] = []
    current_words = 0
    for item in items:
        item = normalize_text(item)
        if not item:
            continue
        words = len(item.split())
        if current and (current_words + words > max_words or len(current) >= max_items):
            chunks.append(current)
            current = [item]
            current_words = words
        else:
            current.append(item)
            current_words += words
    if current:
        chunks.append(current)
    return chunks


def render_article(title: str, fields: Sequence[Tuple[str, str]], bullet_groups: Sequence[Tuple[str, Sequence[str]]] = ()) -> str:
    parts = [f'<article class="guide-record"><h3>{htmlize_text(title)}</h3>']
    for label, value in fields:
        value = normalize_text(value)
        if value:
            parts.append(f'<p><strong>{html.escape(label)}:</strong> {htmlize_text(value)}</p>')
    for label, items in bullet_groups:
        clean_items = [normalize_text(item) for item in items if normalize_text(item)]
        if not clean_items:
            continue
        parts.append(f'<div class="record-list"><p><strong>{html.escape(label)}:</strong></p><ul>')
        for item in clean_items:
            parts.append(f'<li>{htmlize_text(item)}</li>')
        parts.append('</ul></div>')
    parts.append('</article>')
    return ''.join(parts)


def trim_matches_decor(trim: TrimDef, decor_value: str) -> bool:
    decor = normalize_text(decor_value)
    if not decor:
        return False
    parts = re.split(r'\s*/\s*|\s*,\s*', decor)
    trim_values = {trim.name.lower(), trim.code.lower()}
    for part in parts:
        p = normalize_text(part).lower()
        if not p:
            continue
        if p in trim_values:
            return True
        if trim.name.lower() == p:
            return True
        if trim.code.lower().startswith(p) or p.startswith(trim.code.lower()):
            return True
        name_words = trim.name.lower().split()
        if len(name_words) == 1 and p == name_words[0]:
            return True
    return False


def column_matches_trim(column: SpecColumn, trim: TrimDef) -> bool:
    blob = ' '.join([column.top_label, column.header] + column.header_lines).lower()
    if trim.name.lower() in blob or trim.code.lower() in blob:
        return True
    return False


def collect_row_note_texts(row: MatrixRow, sheet: MatrixSheet) -> List[str]:
    notes: List[str] = []
    for note_text in row.inline_footnotes.values():
        notes.append(note_text)
    for raw in row.status_by_trim.values():
        raw = normalize_text(raw)
        if not raw:
            continue
        m = re.match(r'^(--|[A-Z]+|[■□*]+)(.*)$', raw)
        suffix = m.group(2) if m else ''
        for note_id in re.findall(r'\d+', suffix):
            if note_id in row.inline_footnotes:
                notes.append(row.inline_footnotes[note_id])
            elif note_id in sheet.footnotes:
                notes.append(sheet.footnotes[note_id])
    notes.extend(row.bullet_notes)
    return unique_preserve_order([normalize_text(n) for n in notes if normalize_text(n)])


def summarize_model_status_groups(row: MatrixRow, trim_defs: Sequence[TrimDef], sheet: MatrixSheet) -> Tuple[Tuple[str, str, Tuple[str, ...]], ...]:
    groups: 'OrderedDict[Tuple[str, str], List[str]]' = OrderedDict()
    for trim in trim_defs:
        raw = normalize_text(row.status_by_trim.get(trim.key))
        if not raw:
            continue
        _code, label, _notes = parse_status_value(raw, row.inline_footnotes, sheet.footnotes)
        groups.setdefault((raw, label), []).append(trim.name or trim.code)
    return tuple((raw, label, tuple(names)) for (raw, label), names in groups.items())


def model_status_summary_lines(signature: Tuple[Tuple[str, str, Tuple[str, ...]], ...]) -> List[str]:
    lines = []
    for raw, label, names in signature:
        lines.append(f'{label} [{raw}]: {", ".join(names)}')
    return lines


def sort_trim_feature(agg: TrimFeatureAggregate) -> Tuple[int, str]:
    first_key = next(iter(agg.availability_contexts.keys()), ('', ''))
    raw, label = first_key
    return (STATUS_PRIORITY.get(label, 99), normalize_text(agg.title).lower(), raw)


def aggregate_model_features(data: WorkbookData) -> List[ModelFeatureAggregate]:
    groups: 'OrderedDict[str, ModelFeatureAggregate]' = OrderedDict()
    for sheet in data.matrix_sheets:
        for row in sheet.rows:
            title = feature_title(row.label, row.option_code or '', row.ref_code or '')
            agg = groups.get(row.identity_key)
            if agg is None:
                agg = ModelFeatureAggregate(
                    title=title,
                    description=normalize_text(row.description_main or row.description_raw),
                    orderable_code=normalize_text(row.option_code),
                    reference_code=normalize_text(row.ref_code),
                )
                groups[row.identity_key] = agg
            else:
                candidate_desc = normalize_text(row.description_main or row.description_raw)
                if len(candidate_desc) > len(agg.description):
                    agg.description = candidate_desc
                if not agg.orderable_code and row.option_code:
                    agg.orderable_code = normalize_text(row.option_code)
                if not agg.reference_code and row.ref_code:
                    agg.reference_code = normalize_text(row.ref_code)
            ctx = source_context(sheet.name, row.row_group)
            agg.source_contexts = unique_preserve_order(agg.source_contexts + [ctx])
            signature = summarize_model_status_groups(row, sheet.trim_defs, sheet)
            agg.availability_contexts.setdefault(signature, [])
            agg.availability_contexts[signature] = unique_preserve_order(agg.availability_contexts[signature] + [ctx])
            agg.notes = unique_preserve_order(agg.notes + collect_row_note_texts(row, sheet))
            agg.referenced_codes = list(OrderedDict(((code, desc), None) for code, desc in (agg.referenced_codes + referenced_codes_for_text(row.description_raw, data.glossary))).keys())
    return list(groups.values())


def aggregate_trim_features(data: WorkbookData, trim: TrimDef) -> List[TrimFeatureAggregate]:
    groups: 'OrderedDict[str, TrimFeatureAggregate]' = OrderedDict()
    for sheet in data.matrix_sheets:
        for row in sheet.rows:
            raw = normalize_text(row.status_by_trim.get(trim.key))
            if not raw:
                continue
            _code, label, _notes = parse_status_value(raw, row.inline_footnotes, sheet.footnotes)
            title = feature_title(row.label, row.option_code or '', row.ref_code or '')
            agg = groups.get(row.identity_key)
            if agg is None:
                agg = TrimFeatureAggregate(
                    title=title,
                    description=normalize_text(row.description_main or row.description_raw),
                    orderable_code=normalize_text(row.option_code),
                    reference_code=normalize_text(row.ref_code),
                )
                groups[row.identity_key] = agg
            else:
                candidate_desc = normalize_text(row.description_main or row.description_raw)
                if len(candidate_desc) > len(agg.description):
                    agg.description = candidate_desc
                if not agg.orderable_code and row.option_code:
                    agg.orderable_code = normalize_text(row.option_code)
                if not agg.reference_code and row.ref_code:
                    agg.reference_code = normalize_text(row.ref_code)
            ctx = source_context(sheet.name, row.row_group)
            agg.source_contexts = unique_preserve_order(agg.source_contexts + [ctx])
            agg.availability_contexts.setdefault((raw, label), [])
            agg.availability_contexts[(raw, label)] = unique_preserve_order(agg.availability_contexts[(raw, label)] + [ctx])
            agg.notes = unique_preserve_order(agg.notes + collect_row_note_texts(row, sheet))
            agg.referenced_codes = list(OrderedDict(((code, desc), None) for code, desc in (agg.referenced_codes + referenced_codes_for_text(row.description_raw, data.glossary))).keys())
    return sorted(groups.values(), key=sort_trim_feature)


def collect_referenced_codes_for_model(data: WorkbookData) -> List[str]:
    codes: List[str] = []
    for sheet in data.matrix_sheets:
        for row in sheet.rows:
            for code, _desc in referenced_codes_for_text(row.description_raw, data.glossary):
                codes.append(code)
            for code in [row.option_code, row.ref_code]:
                if normalize_text(code) in data.glossary:
                    codes.append(normalize_text(code))
    for sheet in data.color_sheets:
        for row in sheet.interior_rows:
            if normalize_text(row.seat_code) in data.glossary:
                codes.append(normalize_text(row.seat_code))
            for value in row.colors.values():
                if normalize_text(value) in data.glossary:
                    codes.append(normalize_text(value))
        for row in sheet.exterior_rows:
            if normalize_text(row.color_code) in data.glossary:
                codes.append(normalize_text(row.color_code))
    return unique_preserve_order(codes)


def referenced_glossary_codes_for_trim(data: WorkbookData, trim: TrimDef) -> List[str]:
    codes: List[str] = []
    for agg in aggregate_trim_features(data, trim):
        for code in [agg.orderable_code, agg.reference_code]:
            if code and code in data.glossary:
                codes.append(code)
        for code, _desc in agg.referenced_codes:
            if code in data.glossary:
                codes.append(code)
    for sheet in data.color_sheets:
        for row in sheet.interior_rows:
            if not trim_matches_decor(trim, row.decor_level):
                continue
            if row.seat_code and row.seat_code in data.glossary:
                codes.append(row.seat_code)
            for value in row.colors.values():
                if value and value in data.glossary:
                    codes.append(value)
        for row in sheet.exterior_rows:
            if row.color_code and row.color_code in data.glossary:
                codes.append(row.color_code)
    return unique_preserve_order(codes)


def render_note_articles(title_prefix: str, note_texts: Sequence[str], context_text: str, label: str = 'Guide note') -> str:
    parts: List[str] = []
    for note_index, note in enumerate(unique_preserve_order(note_texts), start=1):
        chunks = sentence_chunks(note, max_words=105)
        for chunk_index, chunk in enumerate(chunks, start=1):
            title = f'{title_prefix} | {label.lower()} {note_index}'
            if len(chunks) > 1:
                title += f' | part {chunk_index}'
            parts.append(
                render_article(
                    title,
                    [(label, chunk), ('Applies to source context', context_text)],
                )
            )
    return ''.join(parts)


def render_guide_context_section(data: WorkbookData, page_title: str, trim: Optional[TrimDef] = None) -> str:
    parts = [f'<section class="guide-context"><h2>{html.escape(page_title)} | Guide Context</h2>']
    trim_headers = [trim_def.raw_header for trim_def in data.trim_defs if normalize_text(trim_def.raw_header)]
    fields = [
        ('Source tabs', '; '.join(data.sheet_names)),
        ('Trim headers from guide', ' ; '.join(trim_headers)),
    ]
    if trim is not None:
        fields = [
            ('Trim name', trim.name),
            ('Trim code', trim.code),
            ('Trim header from guide', trim.raw_header),
            ('Source tabs', '; '.join(data.sheet_names)),
        ]
    parts.append(render_article(f'{page_title} | Guide structure', fields))
    parts.append('</section>')
    return ''.join(parts)


def render_matrix_legend_section(data: WorkbookData, page_title: str) -> str:
    if not data.matrix_sheets:
        return ''
    legend_text = normalize_text(data.matrix_sheets[0].legend_text)
    if not legend_text:
        return ''
    return '<section class="matrix-legend">' + render_article(
        f'{page_title} | Matrix availability legend',
        [('Legend from guide', legend_text)],
    ) + '</section>'


def render_model_feature_sections(data: WorkbookData) -> str:
    features = aggregate_model_features(data)
    parts = ['<section class="matrix-features"><h2>Aggregated feature availability from guide</h2>']
    for agg in features:
        description_chunks = sentence_chunks(agg.description, max_words=90)
        fields = []
        if description_chunks:
            fields.append(('Guide text', description_chunks[0]))
        code_bits = [f'Orderable {agg.orderable_code}' for agg in [agg] if agg.orderable_code] + [f'Reference {agg.reference_code}' for agg in [agg] if agg.reference_code]
        if code_bits:
            fields.append(('RPO codes', '; '.join(code_bits)))
        if agg.source_contexts:
            fields.append(('Source context', '; '.join(agg.source_contexts)))
        availability_items: List[str] = []
        for signature, contexts in agg.availability_contexts.items():
            context_text = '; '.join(contexts)
            summary_text = ' ; '.join(model_status_summary_lines(signature))
            if context_text:
                availability_items.append(f'{context_text}: {summary_text}')
            else:
                availability_items.append(summary_text)
        bullet_groups = []
        for availability_chunk in chunk_list(availability_items, max_words=105, max_items=3):
            bullet_groups.append(('Availability by trim', availability_chunk))
        if agg.referenced_codes:
            code_items = [f'{code}: {desc}' for code, desc in agg.referenced_codes]
            for idx, code_chunk in enumerate(chunk_list(code_items, max_words=95, max_items=6), start=1):
                bullet_groups.append((f'Referenced codes from the guide ({idx})' if len(code_items) > len(code_chunk) else 'Referenced codes from the guide', code_chunk))
        parts.append(render_article(agg.title, fields, bullet_groups))
        for idx, extra_desc in enumerate(description_chunks[1:], start=2):
            parts.append(render_article(f'{agg.title} | guide text part {idx - 1}', [('Guide text', extra_desc), ('Source context', '; '.join(agg.source_contexts))]))
        parts.append(render_note_articles(agg.title, agg.notes, '; '.join(agg.source_contexts)))
    parts.append('</section>')
    return ''.join(parts)


def render_model_color_sections(data: WorkbookData) -> str:
    if not data.color_sheets:
        return ''
    parts = ['<section class="colour-trim-sections"><h2>Colour and trim from guide</h2>']
    for sheet in data.color_sheets:
        parts.append(f'<section class="colour-sheet" data-source-tab="{html.escape(sheet.name)}">')
        parts.append(f'<h3>{html.escape(sheet.name)}</h3>')
        for row in sheet.interior_rows:
            color_lines = [f'{color}: {code}' for color, code in row.colors.items() if normalize_text(code) and normalize_text(code) != '--']
            parts.append(
                render_article(
                    feature_title(f'Interior trim | {row.decor_level} | {row.seat_trim}', row.seat_code),
                    [
                        ('Decor level', row.decor_level),
                        ('Seat type', row.seat_type),
                        ('Seat trim', row.seat_trim),
                        ('Source sheet', sheet.name),
                    ],
                    [('Interior colours and guide values', color_lines)] if color_lines else [],
                )
            )
        for row in sheet.exterior_rows:
            availability_lines = [f'{color}: {status}' for color, status in row.colors.items() if normalize_text(status)]
            title_value, title_note_ids = parse_value_and_footnote_ids(row.title)
            note_texts = [sheet.footnotes[nid] for nid in title_note_ids if nid in sheet.footnotes]
            parts.append(
                render_article(
                    feature_title(f'Exterior paint | {title_value or row.title}', row.color_code),
                    [
                        ('Exterior paint', title_value or row.title),
                        ('Touch-Up Paint Number', row.touch_up_paint_number),
                        ('Source sheet', sheet.name),
                    ],
                    [('Availability by interior colour column', availability_lines)] if availability_lines else [],
                )
            )
            parts.append(render_note_articles(feature_title(f'Exterior paint | {title_value or row.title}', row.color_code), note_texts, sheet.name, label='Paint note'))
        note_texts = list(sheet.footnotes.values()) + list(sheet.bullet_notes)
        if note_texts:
            parts.append(render_note_articles(f'{sheet.name} | Colour and trim notes', note_texts, sheet.name, label='Guide note'))
        parts.append('</section>')
    parts.append('</section>')
    return ''.join(parts)


def render_spec_sections(data: WorkbookData, page_title: str, columns: List[SpecColumn]) -> str:
    if not columns:
        return ''
    parts = [f'<section class="spec-sections"><h2>{html.escape(page_title)} | Specifications and dimensions</h2>']
    for column in columns:
        grouped: 'OrderedDict[str, List[str]]' = OrderedDict()
        for cell in column.cells:
            grouped.setdefault(cell.section or 'Data', []).append(f'{cell.label}: {cell.value}')
        header_context = unique_preserve_order([x for x in [column.top_label, column.header] + column.header_lines if normalize_text(x)])
        header_text = ' | '.join(header_context)
        for section_name, values in grouped.items():
            value_chunks = chunk_list(values, max_words=95, max_items=7)
            for idx, value_chunk in enumerate(value_chunks, start=1):
                title = ' | '.join(x for x in [column.header or column.top_label, section_name] if normalize_text(x))
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        title,
                        [('Source sheet', column.sheet_name), ('Column context', header_text)],
                        [('Guide values', value_chunk)],
                    )
                )
    parts.append('</section>')
    return ''.join(parts)


def render_engine_axles_section(data: WorkbookData) -> str:
    if not data.engine_axle_entries:
        return ''
    parts = ['<section class="engine-axles"><h2>Engine, transmission, axle and GVWR from guide</h2>']
    for entry in data.engine_axle_entries:
        grouped: 'OrderedDict[str, List[str]]' = OrderedDict()
        for item in entry.items:
            line = f'{item.name}: {item.status_label} [{item.raw_status}]'
            grouped.setdefault(item.category or 'Guide values', []).append(line)
        for category, items in grouped.items():
            for idx, chunk in enumerate(chunk_list(items, max_words=95, max_items=6), start=1):
                title = ' | '.join(x for x in [entry.model_code, entry.engine, category] if normalize_text(x))
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        title,
                        [('Source sheet', entry.sheet_name), ('Top label', entry.top_label)],
                        [('Guide values', chunk)],
                    )
                )
    parts.append('</section>')
    return ''.join(parts)


def render_trailering_section(data: WorkbookData) -> str:
    if not data.trailering_records and not data.gcwr_records:
        return ''
    parts = ['<section class="trailering"><h2>Trailering and GCWR from guide</h2>']
    for record in data.trailering_records:
        title = ' | '.join(x for x in [record.model_code, record.engine, record.axle_ratio] if normalize_text(x))
        fields = [
            ('Source sheet', record.sheet_name),
            ('Rating type', record.rating_type),
            ('Maximum trailer weight', record.max_trailer_weight),
        ]
        if record.note_text:
            for idx, chunk in enumerate(sentence_chunks(record.note_text, max_words=90), start=1):
                note_title = title + ' | trailering note'
                if idx > 1:
                    note_title += f' | part {idx}'
                parts.append(render_article(note_title, [('Guide text', chunk), ('Source sheet', record.sheet_name)]))
        parts.append(render_article(title, fields, [('Footnotes', record.footnotes)] if record.footnotes else []))
    for record in data.gcwr_records:
        title = ' | '.join(x for x in ['GCWR', record.engine, record.gcwr] if normalize_text(x))
        parts.append(
            render_article(
                title,
                [('Source sheet', record.sheet_name), ('Table title', record.table_title), ('Axle ratio', record.axle_ratio)],
                [('Footnotes', record.footnotes)] if record.footnotes else [('Guide values', [f'GCWR: {record.gcwr}'])],
            )
        )
    parts.append('</section>')
    return ''.join(parts)


def render_glossary_section(page_title: str, glossary: Dict[str, str], limit_codes: Optional[Sequence[str]] = None) -> str:
    if not glossary:
        return ''
    if limit_codes is None:
        ordered_codes = list(glossary.keys())
    else:
        ordered_codes = [code for code in OrderedDict((normalize_text(code), None) for code in limit_codes) if code in glossary]
    if not ordered_codes:
        return ''
    parts = [f'<section class="glossary"><h2>{html.escape(page_title)} | Option code glossary</h2>']
    for code in ordered_codes:
        parts.append(render_article(f'Option code | {code}', [('Option code', code), ('Description from All tab', glossary[code])]))
    parts.append('</section>')
    return ''.join(parts)


def render_trim_feature_sections(data: WorkbookData, trim: TrimDef) -> str:
    features = aggregate_trim_features(data, trim)
    parts = [f'<section class="trim-features"><h2>{html.escape(trim.raw_header or trim.name)} | Features and availability from guide</h2>']
    for agg in features:
        description_chunks = sentence_chunks(agg.description, max_words=90)
        fields = []
        if description_chunks:
            fields.append(('Guide text', description_chunks[0]))
        code_bits = [f'Orderable {agg.orderable_code}' for agg in [agg] if agg.orderable_code] + [f'Reference {agg.reference_code}' for agg in [agg] if agg.reference_code]
        if code_bits:
            fields.append(('RPO codes', '; '.join(code_bits)))
        availability_items = []
        for (raw, label), contexts in agg.availability_contexts.items():
            context_text = '; '.join(contexts)
            if context_text:
                availability_items.append(f'{context_text}: {label} [{raw}]')
            else:
                availability_items.append(f'{label} [{raw}]')
        bullet_groups = [('Source tabs and availability', availability_items)] if availability_items else []
        if agg.referenced_codes:
            code_items = [f'{code}: {desc}' for code, desc in agg.referenced_codes]
            for idx, code_chunk in enumerate(chunk_list(code_items, max_words=95, max_items=6), start=1):
                bullet_groups.append((f'Referenced codes from the guide ({idx})' if len(code_items) > len(code_chunk) else 'Referenced codes from the guide', code_chunk))
        parts.append(render_article(agg.title, fields, bullet_groups))
        for idx, extra_desc in enumerate(description_chunks[1:], start=2):
            parts.append(render_article(f'{agg.title} | guide text part {idx - 1}', [('Guide text', extra_desc), ('Source context', '; '.join(agg.source_contexts))]))
        parts.append(render_note_articles(agg.title, agg.notes, '; '.join(agg.source_contexts)))
    parts.append('</section>')
    return ''.join(parts)


def render_trim_color_sections(data: WorkbookData, trim: TrimDef) -> str:
    if not data.color_sheets:
        return ''
    parts = [f'<section class="trim-colours"><h2>{html.escape(trim.raw_header or trim.name)} | Colour and trim from guide</h2>']
    for sheet in data.color_sheets:
        parts.append(f'<section class="colour-sheet" data-source-tab="{html.escape(sheet.name)}"><h3>{html.escape(sheet.name)}</h3>')
        relevant_interior_columns: List[str] = []
        for row in sheet.interior_rows:
            if not trim_matches_decor(trim, row.decor_level):
                continue
            for color_name, code_value in row.colors.items():
                code_value = normalize_text(code_value)
                if not code_value or code_value == '--':
                    continue
                relevant_interior_columns.append(color_name)
                parts.append(
                    render_article(
                        feature_title(f'Interior colour | {color_name}', row.seat_code),
                        [
                            ('Decor level', row.decor_level),
                            ('Seat type', row.seat_type),
                            ('Seat trim', row.seat_trim),
                            ('Guide value', code_value),
                            ('Source sheet', sheet.name),
                        ],
                    )
                )
        relevant_interior_columns = unique_preserve_order(relevant_interior_columns)
        for row in sheet.exterior_rows:
            availability_lines = []
            for color_name in relevant_interior_columns:
                raw = normalize_text(row.colors.get(color_name))
                if not raw:
                    continue
                _code, label, _notes = parse_status_value(raw, {}, sheet.footnotes)
                availability_lines.append(f'{color_name}: {label} [{raw}]')
            if not availability_lines:
                continue
            title_value, title_note_ids = parse_value_and_footnote_ids(row.title)
            note_texts = [sheet.footnotes[nid] for nid in title_note_ids if nid in sheet.footnotes]
            title = feature_title(f'Exterior paint | {title_value or row.title}', row.color_code)
            parts.append(
                render_article(
                    title,
                    [('Touch-Up Paint Number', row.touch_up_paint_number), ('Source sheet', sheet.name)],
                    [('Availability by interior colour', availability_lines)],
                )
            )
            parts.append(render_note_articles(title, note_texts, sheet.name, label='Paint note'))
        parts.append('</section>')
    parts.append('</section>')
    return ''.join(parts)


def render_trim_spec_sections(data: WorkbookData, trim: TrimDef) -> str:
    direct_columns = [column for column in data.spec_columns if column_matches_trim(column, trim)]
    if direct_columns:
        return render_spec_sections(data, trim.raw_header or trim.name, direct_columns)
    return render_spec_sections(data, trim.raw_header or trim.name, data.spec_columns)


def render_model_page(data: WorkbookData) -> str:
    page_title = 'Vehicle Order Guide'
    referenced_codes = collect_referenced_codes_for_model(data)
    parts = [
        '<html><head><meta charset="utf-8"></head><body>',
        '<h1>Vehicle Order Guide | model page</h1>',
        render_guide_context_section(data, page_title),
        render_matrix_legend_section(data, page_title),
        render_model_feature_sections(data),
        render_model_color_sections(data),
        render_spec_sections(data, page_title, data.spec_columns),
        render_engine_axles_section(data),
        render_trailering_section(data),
        render_glossary_section(page_title, data.glossary, limit_codes=referenced_codes),
        '</body></html>',
    ]
    return ''.join(part for part in parts if part)


def render_trim_page(data: WorkbookData, trim: TrimDef) -> str:
    page_title = trim.raw_header or trim.name
    referenced_codes = referenced_glossary_codes_for_trim(data, trim)
    parts = [
        '<html><head><meta charset="utf-8"></head><body>',
        f'<h1>{html.escape(trim.raw_header or trim.name)} | Vehicle Order Guide trim page</h1>',
        render_guide_context_section(data, page_title, trim=trim),
        render_matrix_legend_section(data, page_title),
        render_trim_feature_sections(data, trim),
        render_trim_color_sections(data, trim),
        render_trim_spec_sections(data, trim),
        render_glossary_section(page_title, data.glossary, limit_codes=referenced_codes),
        '</body></html>',
    ]
    return ''.join(part for part in parts if part)


def write_outputs(data: WorkbookData, output_dir: Path) -> Dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        'workbook': str(data.path),
        'vehicle_name': data.vehicle_name,
        'files': [],
    }

    model_filename = f'model_{data.year}_{slugify(data.make)}_{slugify(data.model)}.html'
    model_path = output_dir / model_filename
    model_path.write_text(render_model_page(data), encoding='utf-8')
    manifest['files'].append({'type': 'model', 'name': data.vehicle_name, 'path': str(model_path)})

    for trim in data.trim_defs:
        trim_filename = f'trim_{data.year}_{slugify(data.make)}_{slugify(data.model)}_{slugify(trim.name)}.html'
        trim_path = output_dir / trim_filename
        trim_path.write_text(render_trim_page(data, trim), encoding='utf-8')
        manifest['files'].append({'type': 'trim', 'name': trim.name, 'path': str(trim_path)})

    manifest_path = output_dir / 'manifest.json'
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding='utf-8')
    return manifest


WORKBOOKS_DIR = Path("workbooks")
OUTPUT_DIR = Path("workbook_html")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description='Convert a GM Vehicle Order Guide workbook into chunk-budget-aware model and trim HTML files for RAG.'
    )
    parser.add_argument(
        'files',
        nargs='*',
        type=Path,
        help='Specific xlsx files to convert. Defaults to all files in workbooks/.',
    )
    parser.add_argument('-o', '--output-dir', help='Directory for generated HTML files')
    args = parser.parse_args(argv)

    output_dir = Path(args.output_dir) if args.output_dir else OUTPUT_DIR

    xlsx_files = args.files if args.files else list(WORKBOOKS_DIR.glob("*.xlsx"))

    if not xlsx_files:
        print(f"No xlsx files found in {WORKBOOKS_DIR}/")
        return 0

    for workbook_path in xlsx_files:
        if not workbook_path.exists():
            print(f'Workbook not found: {workbook_path}', file=sys.stderr)
            continue
        data = parse_workbook(workbook_path)
        manifest = write_outputs(data, output_dir)
        print(json.dumps(manifest, indent=2))
    return 0


# --- Hybrid RAG rendering revision: stronger entity identity, grouped category passages,
# --- atomic feature records, and no glossary output.

CATEGORY_SEQUENCE = [
    'Safety and driver assistance',
    'Technology and connectivity',
    'Interior and comfort',
    'Exterior and utility',
    'Wheels and tires',
    'Mechanical and performance',
    'Packages and options',
    'Colour and trim',
    'Specifications and dimensions',
    'Engine, axle and GVWR',
    'Trailering and GCWR',
    'Other guide content',
]
CATEGORY_ORDER = {name: idx for idx, name in enumerate(CATEGORY_SEQUENCE)}


def page_entity(data: WorkbookData, trim: Optional[TrimDef] = None) -> str:
    if trim is None:
        return data.vehicle_name
    return normalize_text(f'{data.vehicle_name} {trim.name}')


def full_trim_heading(data: WorkbookData, trim: TrimDef) -> str:
    base = page_entity(data, trim)
    if trim.code and trim.code.lower() != trim.name.lower():
        return f'{base} ({trim.code})'
    return base


def article_heading(entity: str, base_title: str) -> str:
    base_title = normalize_text(base_title)
    if not base_title:
        return entity
    return f'{entity} | {base_title}'


def source_tabs_from_contexts(contexts: Sequence[str]) -> str:
    tabs: List[str] = []
    for context in contexts:
        context = normalize_text(context)
        if not context:
            continue
        first = normalize_text(context.split('|', 1)[0])
        if first:
            tabs.append(first)
    return '; '.join(unique_preserve_order(tabs))


def compact_text(text: str, max_words: int = 20) -> str:
    text = normalize_text(text)
    if not text:
        return ''
    words = text.split()
    if len(words) <= max_words:
        return text
    return ' '.join(words[:max_words]).rstrip(',;:') + '...'


def dedupe_fields(fields: Sequence[Tuple[str, str]]) -> List[Tuple[str, str]]:
    cleaned: List[Tuple[str, str]] = []
    seen = set()
    for label, value in fields:
        label = normalize_text(label)
        value = normalize_text(value)
        if not label or not value:
            continue
        key = (label, value)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append((label, value))
    return cleaned


def identity_fields(
    data: WorkbookData,
    trim: Optional[TrimDef] = None,
    *,
    category: str = '',
    source_context: str = '',
    source_tabs: str = '',
    extra_fields: Sequence[Tuple[str, str]] = (),
) -> List[Tuple[str, str]]:
    fields: List[Tuple[str, str]] = [('Vehicle', data.vehicle_name)]
    if trim is not None:
        fields.append(('Trim', trim.name))
        if trim.code:
            fields.append(('Trim code', trim.code))
    if category:
        fields.append(('Guide category', category))
    if source_context:
        fields.append(('Source context', source_context))
    if source_tabs:
        fields.append(('Source tabs', source_tabs))
    fields.extend(extra_fields)
    return dedupe_fields(fields)


def infer_feature_category(*texts: str) -> str:
    blob = ' '.join(normalize_text(text).lower() for text in texts if normalize_text(text))
    if not blob:
        return 'Other guide content'

    if 'colour and trim' in blob or 'color and trim' in blob or ' paint' in blob or blob.startswith('paint ') or 'decor level' in blob:
        return 'Colour and trim'

    safety_keywords = [
        'airbag', 'airbags', 'blind zone', 'collision', 'cruise', 'driver assistance', 'following distance',
        'hd surround vision', 'lane ', 'pedestrian', 'parking assist', 'rear cross traffic', 'rear pedestrian',
        'safety', 'seat belt', 'stability control', 'traffic sign', 'traction control', 'warning', 'alert',
        'camera', 'restraint', 'teen driver', 'sensing system', 'automatic emergency braking', 'brake assist',
        'reverse automatic braking', 'tire pressure monitor', 'rear park assist'
    ]
    if any(keyword in blob for keyword in safety_keywords):
        return 'Safety and driver assistance'

    technology_keywords = [
        'android auto', 'apple carplay', 'audio system', 'bluetooth', 'display', 'google built-in',
        'head-up display', 'infotainment', 'mychevrolet', 'navigation', 'onstar', 'phone', 'radio',
        'remote start', 'screen', 'siriusxm', 'smartphone', 'speaker', 'usb', 'wi-fi', 'wifi',
        'wireless', 'charging pad', 'charging-only', 'device charging', 'driver information center'
    ]
    if any(keyword in blob for keyword in technology_keywords):
        return 'Technology and connectivity'

    wheels_keywords = ['wheel', 'wheels', 'tire', 'tires', 'spare tire', 'spare wheel', 'lug nut', 'wheel lock']
    if any(keyword in blob for keyword in wheels_keywords):
        return 'Wheels and tires'

    mechanical_keywords = [
        'all-wheel drive', 'axle', 'battery', 'brakes', 'charging', 'charger', 'drive unit', 'drivetrain',
        'electric drive', 'engine', 'evot? ', 'fuel', 'gvwr', 'horsepower', 'motor', 'payload', 'performance',
        'powertrain', 'propulsion', 'range', 'rear axle', 'suspension', 'torque', 'tow', 'trailer',
        'trailering', 'transmission'
    ]
    if any(keyword in blob for keyword in mechanical_keywords):
        return 'Mechanical and performance'

    exterior_keywords = [
        'bed', 'box', 'bumper', 'cargo', 'cross rails', 'door', 'emblem', 'fascia', 'glass', 'grille',
        'headlamp', 'hood', 'lamp', 'liftgate', 'license plate', 'mirror, outside', 'nameplate', 'roof',
        'running board', 'splash guard', 'tailgate', 'window', 'wiper', 'privacy glass', 'deep tint'
    ]
    if any(keyword in blob for keyword in exterior_keywords):
        return 'Exterior and utility'

    interior_keywords = [
        'air conditioning', 'ambient', 'armrest', 'carpet', 'climate', 'console', 'cup holder', 'driver seat',
        'floor mat', 'headrest', 'heated seat', 'inside rearview', 'instrument panel', 'lumbar', 'rear seat',
        'seat adjuster', 'seat trim', 'seating', 'seats', 'steering wheel', 'sun visor', 'visor', 'interior'
    ]
    if any(keyword in blob for keyword in interior_keywords):
        return 'Interior and comfort'

    package_keywords = ['package', 'equipment group', 'lpo', 'accessory', 'dealer-installed', 'option']
    if any(keyword in blob for keyword in package_keywords):
        return 'Packages and options'

    if 'interior' in blob:
        return 'Interior and comfort'
    if 'exterior' in blob:
        return 'Exterior and utility'
    if 'mechanical' in blob:
        return 'Mechanical and performance'
    if 'wheels' in blob:
        return 'Wheels and tires'
    if 'onstar' in blob or 'siriusxm' in blob:
        return 'Technology and connectivity'
    return 'Other guide content'


def sort_category_key(category: str) -> Tuple[int, str]:
    return (CATEGORY_ORDER.get(category, 999), normalize_text(category).lower())


def category_for_trim_feature(agg: TrimFeatureAggregate) -> str:
    return infer_feature_category(' '.join(agg.source_contexts), agg.title, agg.description)


def category_for_model_feature(agg: ModelFeatureAggregate) -> str:
    return infer_feature_category(' '.join(agg.source_contexts), agg.title, agg.description)


def availability_lines_for_trim(agg: TrimFeatureAggregate) -> List[str]:
    lines: List[str] = []
    for (raw, label), contexts in agg.availability_contexts.items():
        context_text = '; '.join(contexts)
        if context_text:
            lines.append(f'{label} [{raw}] — {context_text}')
        else:
            lines.append(f'{label} [{raw}]')
    return lines


def availability_summary_for_trim(agg: TrimFeatureAggregate) -> str:
    parts: List[str] = []
    for (raw, label), contexts in agg.availability_contexts.items():
        tabs = source_tabs_from_contexts(contexts)
        bit = f'{label} [{raw}]'
        if tabs:
            bit += f' ({tabs})'
        parts.append(bit)
    return '; '.join(parts)


def availability_lines_for_model(agg: ModelFeatureAggregate) -> List[str]:
    lines: List[str] = []
    for signature, contexts in agg.availability_contexts.items():
        summary = ' ; '.join(model_status_summary_lines(signature))
        context_text = '; '.join(contexts)
        if context_text:
            lines.append(f'{summary} — {context_text}')
        else:
            lines.append(summary)
    return lines


def availability_summary_for_model(agg: ModelFeatureAggregate) -> str:
    parts: List[str] = []
    for signature, _contexts in agg.availability_contexts.items():
        parts.append(' ; '.join(model_status_summary_lines(signature)))
    return ' / '.join(parts)


def chunk_feature_items(items: Sequence[str]) -> List[List[str]]:
    return chunk_list(items, max_words=135, max_items=8)


def render_page_identity_section(data: WorkbookData, trim: Optional[TrimDef] = None) -> str:
    entity = full_trim_heading(data, trim) if trim is not None else data.vehicle_name
    parts = [f'<section class="guide-context"><h2>{html.escape(entity)} | Vehicle identity and guide structure</h2>']
    trim_headers = [trim_def.raw_header for trim_def in data.trim_defs if normalize_text(trim_def.raw_header)]
    if trim is None:
        fields = [
            ('Vehicle', data.vehicle_name),
            ('Source tabs', '; '.join(data.sheet_names)),
            ('Trim headers from guide', ' ; '.join(trim_headers)),
        ]
        parts.append(render_article(article_heading(entity, 'Vehicle identity and guide structure'), dedupe_fields(fields)))
    else:
        fields = [
            ('Vehicle', data.vehicle_name),
            ('Trim', trim.name),
            ('Trim code', trim.code),
            ('Trim header from guide', trim.raw_header),
            ('Source tabs', '; '.join(data.sheet_names)),
        ]
        parts.append(render_article(article_heading(entity, 'Vehicle identity and guide structure'), dedupe_fields(fields)))
    parts.append('</section>')
    return ''.join(parts)


def render_matrix_legend_section(data: WorkbookData, page_title: str) -> str:
    if not data.matrix_sheets:
        return ''
    legend_text = normalize_text(data.matrix_sheets[0].legend_text)
    if not legend_text:
        return ''
    return '<section class="matrix-legend">' + render_article(
        article_heading(page_title, 'Matrix availability legend'),
        identity_fields(data, category='Other guide content', extra_fields=[('Legend from guide', legend_text)]),
    ) + '</section>'


def trim_group_line(agg: TrimFeatureAggregate) -> str:
    feature_text = compact_text(agg.description or agg.title, max_words=18)
    return f'{feature_text} — {availability_summary_for_trim(agg)}'


def model_group_line(agg: ModelFeatureAggregate) -> str:
    feature_text = compact_text(agg.description or agg.title, max_words=18)
    return f'{feature_text} — {availability_summary_for_model(agg)}'


def grouped_feature_sections(
    data: WorkbookData,
    features: Sequence[object],
    *,
    trim: Optional[TrimDef] = None,
    model_mode: bool = False,
) -> str:
    if not features:
        return ''
    entity = full_trim_heading(data, trim) if trim is not None else data.vehicle_name
    category_groups: Dict[str, List[object]] = OrderedDict()
    for feature in features:
        category = category_for_model_feature(feature) if model_mode else category_for_trim_feature(feature)
        category_groups.setdefault(category, []).append(feature)

    parts = [f'<section class="grouped-feature-passages"><h2>{html.escape(entity)} | Grouped feature passages from guide</h2>']
    for category in sorted(category_groups.keys(), key=sort_category_key):
        items = category_groups[category]
        lines = [model_group_line(item) if model_mode else trim_group_line(item) for item in items]
        source_tabs = source_tabs_from_contexts([ctx for item in items for ctx in getattr(item, 'source_contexts', [])])
        for idx, line_chunk in enumerate(chunk_feature_items(lines), start=1):
            title = f'{category} | grouped guide passage'
            if idx > 1:
                title += f' | part {idx}'
            parts.append(
                render_article(
                    article_heading(entity, title),
                    identity_fields(data, trim, category=category, source_tabs=source_tabs),
                    [('Feature lines from guide', line_chunk)],
                )
            )
    parts.append('</section>')
    return ''.join(parts)


def exact_model_feature_section(data: WorkbookData, features: Sequence[ModelFeatureAggregate]) -> str:
    if not features:
        return ''
    entity = data.vehicle_name
    parts = [f'<section class="exact-feature-records"><h2>{html.escape(entity)} | Exact feature records from guide</h2>']
    ordered = sorted(features, key=lambda agg: (sort_category_key(category_for_model_feature(agg)), normalize_text(agg.title).lower()))
    for agg in ordered:
        category = category_for_model_feature(agg)
        fields = identity_fields(
            data,
            category=category,
            source_context='; '.join(agg.source_contexts),
        )
        fields.append(('Guide text', agg.description))
        bullet_groups: List[Tuple[str, Sequence[str]]] = []
        availability_lines = availability_lines_for_model(agg)
        if availability_lines:
            bullet_groups.append(('Availability by trim', availability_lines))
        if agg.notes:
            bullet_groups.append(('Guide notes', agg.notes))
        parts.append(render_article(article_heading(entity, agg.title), fields, bullet_groups))
    parts.append('</section>')
    return ''.join(parts)


def exact_trim_feature_section(data: WorkbookData, trim: TrimDef, features: Sequence[TrimFeatureAggregate]) -> str:
    if not features:
        return ''
    entity = full_trim_heading(data, trim)
    parts = [f'<section class="exact-feature-records"><h2>{html.escape(entity)} | Exact feature records from guide</h2>']
    ordered = sorted(features, key=lambda agg: (sort_category_key(category_for_trim_feature(agg)), sort_trim_feature(agg)))
    for agg in ordered:
        category = category_for_trim_feature(agg)
        fields = identity_fields(
            data,
            trim,
            category=category,
            source_context='; '.join(agg.source_contexts),
        )
        fields.append(('Guide text', agg.description))
        bullet_groups: List[Tuple[str, Sequence[str]]] = []
        availability_lines = availability_lines_for_trim(agg)
        if availability_lines:
            bullet_groups.append(('Availability on this trim', availability_lines))
        if agg.notes:
            bullet_groups.append(('Guide notes', agg.notes))
        parts.append(render_article(article_heading(entity, agg.title), fields, bullet_groups))
    parts.append('</section>')
    return ''.join(parts)


def render_model_feature_sections(data: WorkbookData) -> str:
    features = aggregate_model_features(data)
    return ''.join([
        grouped_feature_sections(data, features, model_mode=True),
        exact_model_feature_section(data, features),
    ])


def render_trim_feature_sections(data: WorkbookData, trim: TrimDef) -> str:
    features = aggregate_trim_features(data, trim)
    return ''.join([
        grouped_feature_sections(data, features, trim=trim, model_mode=False),
        exact_trim_feature_section(data, trim, features),
    ])


def render_model_color_sections(data: WorkbookData) -> str:
    if not data.color_sheets:
        return ''
    entity = data.vehicle_name
    parts = [f'<section class="colour-trim-sections"><h2>{html.escape(entity)} | Colour and trim from guide</h2>']
    for sheet in data.color_sheets:
        interior_lines: List[str] = []
        for row in sheet.interior_rows:
            color_lines = [f'{color}: {code}' for color, code in row.colors.items() if normalize_text(code) and normalize_text(code) != '--']
            summary = ' | '.join(x for x in [row.decor_level, row.seat_type, row.seat_trim] if normalize_text(x))
            if color_lines:
                summary += ' — ' + '; '.join(color_lines)
            interior_lines.append(summary)
        if interior_lines:
            for idx, chunk in enumerate(chunk_feature_items(interior_lines), start=1):
                title = 'Colour and trim | interior grouped passage'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(data, category='Colour and trim', source_tabs=sheet.name),
                        [('Interior colour and trim lines from guide', chunk)],
                    )
                )
        exterior_lines: List[str] = []
        for row in sheet.exterior_rows:
            title_value, _title_note_ids = parse_value_and_footnote_ids(row.title)
            availability = [f'{color}: {status}' for color, status in row.colors.items() if normalize_text(status)]
            line = ' | '.join(x for x in [title_value or row.title, row.color_code] if normalize_text(x))
            if availability:
                line += ' — ' + '; '.join(availability)
            exterior_lines.append(line)
        if exterior_lines:
            for idx, chunk in enumerate(chunk_feature_items(exterior_lines), start=1):
                title = 'Colour and trim | exterior paint grouped passage'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(data, category='Colour and trim', source_tabs=sheet.name),
                        [('Exterior paint lines from guide', chunk)],
                    )
                )
        for row in sheet.interior_rows:
            color_lines = [f'{color}: {code}' for color, code in row.colors.items() if normalize_text(code) and normalize_text(code) != '--']
            parts.append(
                render_article(
                    article_heading(entity, feature_title(f'Interior trim | {row.decor_level} | {row.seat_trim}', row.seat_code)),
                    identity_fields(
                        data,
                        category='Colour and trim',
                        source_tabs=sheet.name,
                        extra_fields=[
                            ('Decor level', row.decor_level),
                            ('Seat type', row.seat_type),
                            ('Seat trim', row.seat_trim),
                        ],
                    ),
                    [('Interior colours and guide values', color_lines)] if color_lines else [],
                )
            )
        for row in sheet.exterior_rows:
            availability_lines = [f'{color}: {status}' for color, status in row.colors.items() if normalize_text(status)]
            title_value, title_note_ids = parse_value_and_footnote_ids(row.title)
            note_texts = [sheet.footnotes[nid] for nid in title_note_ids if nid in sheet.footnotes]
            bullet_groups: List[Tuple[str, Sequence[str]]] = []
            if availability_lines:
                bullet_groups.append(('Availability by interior colour column', availability_lines))
            if note_texts:
                bullet_groups.append(('Guide notes', note_texts))
            parts.append(
                render_article(
                    article_heading(entity, feature_title(f'Exterior paint | {title_value or row.title}', row.color_code)),
                    identity_fields(
                        data,
                        category='Colour and trim',
                        source_tabs=sheet.name,
                        extra_fields=[('Touch-Up Paint Number', row.touch_up_paint_number)],
                    ),
                    bullet_groups,
                )
            )
        general_notes = unique_preserve_order(list(sheet.footnotes.values()) + list(sheet.bullet_notes))
        if general_notes:
            parts.append(
                render_article(
                    article_heading(entity, f'{sheet.name} | colour and trim notes'),
                    identity_fields(data, category='Colour and trim', source_tabs=sheet.name),
                    [('Guide notes', general_notes)],
                )
            )
    parts.append('</section>')
    return ''.join(parts)


def render_trim_color_sections(data: WorkbookData, trim: TrimDef) -> str:
    if not data.color_sheets:
        return ''
    entity = full_trim_heading(data, trim)
    parts = [f'<section class="trim-colours"><h2>{html.escape(entity)} | Colour and trim from guide</h2>']
    for sheet in data.color_sheets:
        grouped_interior_lines: List[str] = []
        relevant_interior_columns: List[str] = []
        for row in sheet.interior_rows:
            if not trim_matches_decor(trim, row.decor_level):
                continue
            color_lines = [f'{color}: {code}' for color, code in row.colors.items() if normalize_text(code) and normalize_text(code) != '--']
            relevant_interior_columns.extend([color for color, code in row.colors.items() if normalize_text(code) and normalize_text(code) != '--'])
            summary = ' | '.join(x for x in [row.decor_level, row.seat_type, row.seat_trim] if normalize_text(x))
            if color_lines:
                summary += ' — ' + '; '.join(color_lines)
            grouped_interior_lines.append(summary)
            parts.append(
                render_article(
                    article_heading(entity, feature_title(f'Interior trim | {row.decor_level} | {row.seat_trim}', row.seat_code)),
                    identity_fields(
                        data,
                        trim,
                        category='Colour and trim',
                        source_tabs=sheet.name,
                        extra_fields=[
                            ('Decor level', row.decor_level),
                            ('Seat type', row.seat_type),
                            ('Seat trim', row.seat_trim),
                        ],
                    ),
                    [('Interior colours and guide values', color_lines)] if color_lines else [],
                )
            )
        relevant_interior_columns = unique_preserve_order(relevant_interior_columns)
        if grouped_interior_lines:
            for idx, chunk in enumerate(chunk_feature_items(grouped_interior_lines), start=1):
                title = 'Colour and trim | interior grouped passage'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(data, trim, category='Colour and trim', source_tabs=sheet.name),
                        [('Interior colour and trim lines from guide', chunk)],
                    )
                )
        grouped_exterior_lines: List[str] = []
        for row in sheet.exterior_rows:
            availability_lines: List[str] = []
            for color_name in relevant_interior_columns:
                raw = normalize_text(row.colors.get(color_name))
                if not raw:
                    continue
                _code, label, _notes = parse_status_value(raw, {}, sheet.footnotes)
                availability_lines.append(f'{color_name}: {label} [{raw}]')
            if not availability_lines:
                continue
            title_value, title_note_ids = parse_value_and_footnote_ids(row.title)
            note_texts = [sheet.footnotes[nid] for nid in title_note_ids if nid in sheet.footnotes]
            grouped_line = ' | '.join(x for x in [title_value or row.title, row.color_code] if normalize_text(x))
            grouped_line += ' — ' + '; '.join(availability_lines)
            grouped_exterior_lines.append(grouped_line)
            bullet_groups: List[Tuple[str, Sequence[str]]] = [('Availability by interior colour', availability_lines)]
            if note_texts:
                bullet_groups.append(('Guide notes', note_texts))
            parts.append(
                render_article(
                    article_heading(entity, feature_title(f'Exterior paint | {title_value or row.title}', row.color_code)),
                    identity_fields(
                        data,
                        trim,
                        category='Colour and trim',
                        source_tabs=sheet.name,
                        extra_fields=[('Touch-Up Paint Number', row.touch_up_paint_number)],
                    ),
                    bullet_groups,
                )
            )
        if grouped_exterior_lines:
            for idx, chunk in enumerate(chunk_feature_items(grouped_exterior_lines), start=1):
                title = 'Colour and trim | exterior paint grouped passage'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(data, trim, category='Colour and trim', source_tabs=sheet.name),
                        [('Exterior paint lines from guide', chunk)],
                    )
                )
        general_notes = unique_preserve_order(list(sheet.footnotes.values()) + list(sheet.bullet_notes))
        if general_notes:
            parts.append(
                render_article(
                    article_heading(entity, f'{sheet.name} | colour and trim notes'),
                    identity_fields(data, trim, category='Colour and trim', source_tabs=sheet.name),
                    [('Guide notes', general_notes)],
                )
            )
    parts.append('</section>')
    return ''.join(parts)


def render_spec_records(data: WorkbookData, columns: List[SpecColumn], *, trim: Optional[TrimDef] = None) -> str:
    if not columns:
        return ''
    entity = full_trim_heading(data, trim) if trim is not None else data.vehicle_name
    parts = [f'<section class="spec-sections"><h2>{html.escape(entity)} | Specifications and dimensions from guide</h2>']
    for column in columns:
        grouped: 'OrderedDict[str, List[str]]' = OrderedDict()
        for cell in column.cells:
            grouped.setdefault(cell.section or 'Data', []).append(f'{cell.label}: {cell.value}')
        header_context = unique_preserve_order([x for x in [column.top_label, column.header] + column.header_lines if normalize_text(x)])
        header_text = ' | '.join(header_context)
        for section_name, values in grouped.items():
            for idx, value_chunk in enumerate(chunk_list(values, max_words=110, max_items=8), start=1):
                title = f'Specifications and dimensions | {section_name}'
                if header_text:
                    title += f' | {column.header or column.top_label}'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(
                            data,
                            trim,
                            category='Specifications and dimensions',
                            source_tabs=column.sheet_name,
                            extra_fields=[('Column context', header_text)],
                        ),
                        [('Guide values', value_chunk)],
                    )
                )
    parts.append('</section>')
    return ''.join(parts)


def render_trim_spec_sections(data: WorkbookData, trim: TrimDef) -> str:
    direct_columns = [column for column in data.spec_columns if column_matches_trim(column, trim)]
    if direct_columns:
        return render_spec_records(data, direct_columns, trim=trim)
    return render_spec_records(data, data.spec_columns, trim=trim)


def render_engine_axles_section(data: WorkbookData) -> str:
    if not data.engine_axle_entries:
        return ''
    entity = data.vehicle_name
    parts = [f'<section class="engine-axles"><h2>{html.escape(entity)} | Engine, axle and GVWR from guide</h2>']
    for entry in data.engine_axle_entries:
        grouped: 'OrderedDict[str, List[str]]' = OrderedDict()
        for item in entry.items:
            line = f'{item.name}: {item.status_label} [{item.raw_status}]'
            if item.notes:
                line += ' — ' + '; '.join(item.notes)
            grouped.setdefault(item.category or 'Guide values', []).append(line)
        for category, items in grouped.items():
            for idx, chunk in enumerate(chunk_list(items, max_words=110, max_items=8), start=1):
                title = f'Engine, axle and GVWR | {entry.model_code} | {entry.engine} | {category}'
                if idx > 1:
                    title += f' | part {idx}'
                parts.append(
                    render_article(
                        article_heading(entity, title),
                        identity_fields(
                            data,
                            category='Engine, axle and GVWR',
                            source_tabs=entry.sheet_name,
                            extra_fields=[('Top label', entry.top_label)],
                        ),
                        [('Guide values', chunk)],
                    )
                )
    parts.append('</section>')
    return ''.join(parts)


def render_trailering_section(data: WorkbookData) -> str:
    if not data.trailering_records and not data.gcwr_records:
        return ''
    entity = data.vehicle_name
    parts = [f'<section class="trailering"><h2>{html.escape(entity)} | Trailering and GCWR from guide</h2>']
    for record in data.trailering_records:
        bullet_groups: List[Tuple[str, Sequence[str]]] = []
        if record.note_text:
            bullet_groups.append(('Guide text', sentence_chunks(record.note_text, max_words=90)))
        if record.footnotes:
            bullet_groups.append(('Guide notes', record.footnotes))
        parts.append(
            render_article(
                article_heading(entity, f'Trailering and GCWR | {record.model_code} | {record.engine} | {record.axle_ratio}'),
                identity_fields(
                    data,
                    category='Trailering and GCWR',
                    source_tabs=record.sheet_name,
                    extra_fields=[
                        ('Rating type', record.rating_type),
                        ('Maximum trailer weight', record.max_trailer_weight),
                    ],
                ),
                bullet_groups,
            )
        )
    for record in data.gcwr_records:
        bullet_groups: List[Tuple[str, Sequence[str]]] = []
        if record.footnotes:
            bullet_groups.append(('Guide notes', record.footnotes))
        else:
            bullet_groups.append(('Guide values', [f'GCWR: {record.gcwr}']))
        parts.append(
            render_article(
                article_heading(entity, f'Trailering and GCWR | GCWR | {record.engine} | {record.gcwr}'),
                identity_fields(
                    data,
                    category='Trailering and GCWR',
                    source_tabs=record.sheet_name,
                    extra_fields=[('Table title', record.table_title), ('Axle ratio', record.axle_ratio)],
                ),
                bullet_groups,
            )
        )
    parts.append('</section>')
    return ''.join(parts)


def render_model_page(data: WorkbookData) -> str:
    entity = data.vehicle_name
    parts = [
        '<html><head><meta charset="utf-8"></head><body>',
        f'<h1>{html.escape(entity)} | Vehicle Order Guide model page</h1>',
        render_page_identity_section(data),
        render_matrix_legend_section(data, entity),
        render_model_feature_sections(data),
        render_model_color_sections(data),
        render_spec_records(data, data.spec_columns),
        render_engine_axles_section(data),
        render_trailering_section(data),
        '</body></html>',
    ]
    return ''.join(part for part in parts if part)


def render_trim_page(data: WorkbookData, trim: TrimDef) -> str:
    entity = full_trim_heading(data, trim)
    parts = [
        '<html><head><meta charset="utf-8"></head><body>',
        f'<h1>{html.escape(entity)} | Vehicle Order Guide trim page</h1>',
        render_page_identity_section(data, trim=trim),
        render_matrix_legend_section(data, entity),
        render_trim_feature_sections(data, trim),
        render_trim_color_sections(data, trim),
        render_trim_spec_sections(data, trim),
        '</body></html>',
    ]
    return ''.join(part for part in parts if part)

if __name__ == '__main__':
    raise SystemExit(main())
